"""
control_simple_route.py — Control cruzado Simple Route vs BO San Bernardo (RM)

Pregunta de control:
    "Todo lo que Simple Route dice que fue DF, ¿existe en BO San Bernardo
     con el mismo cliente, contrato y pesaje?"

Es el único control que contrasta las planillas de Ambipar contra un sistema
externo. Los otros tres scripts de RM comparan las fuentes entre sí.

QUÉ SE CAMBIÓ AL REFACTORIZAR (y qué NO)
    Se sacaron las rutas fijas del computador (la carpeta de OneDrive) y se
    expuso controlar(), que recibe las rutas por parámetro. Eso es todo.

    NINGUNA función de cálculo fue modificada: parsear_titulo, es_df,
    seleccionar_peso_sr, preparar_simple_route, preparar_bo, comparar_fila,
    comparar, resumen y estilizar son exactamente las mismas del original.
    El resultado tiene que ser idéntico.

    Se conserva el uso desde la terminal:
        python control_simple_route.py "simple_route_rm.xlsx" --bo "..."

Salida: un Excel con tres hojas
    1) RESUMEN            — la pregunta, los filtros y el conteo por resultado
    2) COMPARACION_DF     — fila por fila, qué encontró en BO y qué no
    3) TITULOS_A_REVISAR  — títulos de Simple Route que no se pudieron separar
"""

import argparse
import re
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Las rutas ya no viven acá: las entrega quien llama a controlar(). En la app
# es la carpeta de trabajo del navegador; en la terminal, los argumentos.
NOMBRE_SALIDA = "CONTROL_SIMPLE_ROUTE_DF_BO_RM.xlsx"

TOLERANCIA_KG = 10
TOLERANCIA_PCT = 0.01

COLUMNAS_PESO_SR_PRIORIDAD = [
    "Pesaje en kilogramos",
    "Peso disposición final (kilos que indica ticket)",
    "Peso total de residuos retirados (En Kilos, en caso de tener pesaje en cliente)",
    "Kilos de material entregados (Kg) Seleccionar 0 si no tiene pesaje.",
    "Peso residuos a valorizar (Kilos, Opcional en caso de tener)",
    "Peso Pallet con fardo (Kilos, Opcional en caso de tener)",
]


class Registro:
    """Guarda los mensajes además de imprimirlos, para poder devolverlos."""

    def __init__(self, mostrar=True):
        self.lineas = []
        self.mostrar = mostrar

    def __call__(self, msg=""):
        self.lineas.append(str(msg))
        if self.mostrar:
            print(msg)

    def texto(self):
        return "\n".join(self.lineas)


def limpiar_texto(x):
    if pd.isna(x):
        return ""
    x = str(x).strip()
    if x.lower() in {"", "nan", "none", "nat"}:
        return ""
    return re.sub(r"\s+", " ", x)


def normalizar_texto(x):
    x = limpiar_texto(x).upper()
    x = unicodedata.normalize("NFKD", x).encode("ascii", "ignore").decode("utf-8")
    x = x.replace(".", "")
    x = re.sub(r"[^A-Z0-9 ]", " ", x)
    x = re.sub(r"\s+", " ", x)
    return x.strip()


def normalizar_contrato(x):
    x = limpiar_texto(x)
    x = re.sub(r"\.0$", "", x)
    x = re.sub(r"[^0-9A-Za-z]", "", x)
    return x.upper().strip()


def normalizar_numero(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "-"}:
        return np.nan
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
        s = s.replace(".", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", ".", "-", "-."}:
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan


def normalizar_fecha(x):
    return pd.to_datetime(x, errors="coerce").dt.normalize()


def pesos_equivalentes(peso_sr, peso_bo):
    if pd.isna(peso_sr) or pd.isna(peso_bo):
        return False
    diferencia = abs(float(peso_sr) - float(peso_bo))
    base = max(abs(float(peso_sr)), 1)
    return diferencia <= max(TOLERANCIA_KG, base * TOLERANCIA_PCT)


def buscar_columna(df, opciones, obligatorio=True):
    mapa = {normalizar_texto(c): c for c in df.columns}
    for op in opciones:
        key = normalizar_texto(op)
        if key in mapa:
            return mapa[key]
    if obligatorio:
        raise ValueError(f"No encontré ninguna de estas columnas: {opciones}")
    return None


def leer_tabla_nombrada(path, nombre_tabla):
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        for sheet in wb.worksheets:
            for tabla in sheet.tables.values():
                if tabla.name == nombre_tabla:
                    filas = [[c.value for c in fila] for fila in sheet[tabla.ref]]
                    headers = filas[0]
                    return pd.DataFrame(filas[1:], columns=headers).dropna(how="all").reset_index(drop=True)
    finally:
        wb.close()
    raise ValueError(f"No se encontró la tabla {nombre_tabla}")


def parsear_titulo(titulo):
    titulo_original = limpiar_texto(titulo)
    if not titulo_original:
        return pd.Series(["", "", "", "Título vacío"])
    partes = [p.strip() for p in re.split(r"\s+-\s+", titulo_original) if p.strip()]
    if len(partes) >= 3 and re.match(r"^\d+", partes[0]):
        contrato = normalizar_contrato(partes[0])
        tipo = partes[1]
        cliente = " - ".join(partes[2:]).strip()
        return pd.Series([contrato, tipo, cliente, "OK"])
    if len(partes) >= 2 and re.match(r"^\d+", partes[0]):
        contrato = normalizar_contrato(partes[0])
        return pd.Series([contrato, "", " - ".join(partes[1:]).strip(), "Falta Tipo/DF en título"])
    return pd.Series(["", "", titulo_original, "Título no separable"])


def es_df(tipo):
    t = normalizar_texto(tipo)
    return bool(t == "D" or "DF" in t or "DISPOSICION FINAL" in t)


def seleccionar_peso_sr(sr):
    existentes = [c for c in COLUMNAS_PESO_SR_PRIORIDAD if c in sr.columns]
    sr["Pesaje_SR"] = np.nan
    sr["Columna_pesaje_SR"] = ""
    for col in existentes:
        valores = sr[col].apply(normalizar_numero)
        mask = sr["Pesaje_SR"].isna() & valores.notna() & (valores != 0)
        sr.loc[mask, "Pesaje_SR"] = valores.loc[mask]
        sr.loc[mask, "Columna_pesaje_SR"] = col
    return sr


def preparar_simple_route(path_sr, comparar_todos_los_estados=False):
    sr = pd.read_excel(path_sr, sheet_name=0, dtype=str)
    sr.columns = sr.columns.astype(str).str.strip()

    col_titulo = buscar_columna(sr, ["Título", "Titulo"])
    col_fecha = buscar_columna(sr, ["Fecha planificada", "Fecha"])
    col_estado = buscar_columna(sr, ["Estado"], obligatorio=False)

    parsed = sr[col_titulo].apply(parsear_titulo)
    parsed.columns = ["Contrato_SR", "Tipo_SR", "Cliente_SR", "Estado_titulo"]
    sr = pd.concat([sr, parsed], axis=1)
    sr = seleccionar_peso_sr(sr)

    sr["Fecha_SR"] = pd.to_datetime(sr[col_fecha], errors="coerce").dt.date
    sr["Fecha_key"] = pd.to_datetime(sr[col_fecha], errors="coerce").dt.normalize()
    sr["Estado_SR"] = sr[col_estado].apply(limpiar_texto) if col_estado else ""
    sr["Estado_key"] = sr["Estado_SR"].apply(normalizar_texto)
    sr["Cliente_key"] = sr["Cliente_SR"].apply(normalizar_texto)
    sr["Contrato_key"] = sr["Contrato_SR"].apply(normalizar_contrato)
    sr["Es_DF"] = sr["Tipo_SR"].apply(es_df)

    if comparar_todos_los_estados:
        comparar = sr["Es_DF"]
    else:
        comparar = sr["Es_DF"] & sr["Estado_key"].eq("COMPLETED")

    columnas = [
        "ID de visita",
        "Tracking ID",
        "Fecha_SR",
        "Estado_SR",
        "Título",
        "Contrato_SR",
        "Tipo_SR",
        "Cliente_SR",
        "Pesaje_SR",
        "Columna_pesaje_SR",
        "Estado_titulo",
        "Es_DF",
        "Fecha_key",
        "Cliente_key",
        "Contrato_key",
    ]
    columnas = [c for c in columnas if c in sr.columns]
    sr_limpia = sr[columnas].copy()
    df_comparar = sr_limpia[comparar].copy().reset_index(drop=True)
    titulos_revisar = sr_limpia[(sr_limpia["Es_DF"]) & (sr_limpia["Estado_titulo"] != "OK")].copy()
    return sr_limpia, df_comparar, titulos_revisar


def preparar_bo(path_bo):
    try:
        bo = leer_tabla_nombrada(path_bo, "Tabla_operacion")
    except Exception:
        bo = pd.read_excel(path_bo, sheet_name=0, dtype=str)
    bo.columns = bo.columns.astype(str).str.strip()

    col_fecha = buscar_columna(bo, ["Fecha", "FECHA"])
    col_cliente = buscar_columna(bo, ["Cliente", "CLIENTE", "Razón Social", "Raz_social"])
    col_contrato = buscar_columna(bo, ["Contrato", "N° CONTRATO", "Nº CONTRATO", "Nro Contrato", "Contratato"])
    col_peso = buscar_columna(bo, ["Peso neto (kg)", "Peso neto", "PESO NETO KG", "Peso Final [kg]"])
    col_ticket = buscar_columna(bo, ["Ticket de pesaje", "TICKET DE PESAJE", "RtHruFol", "Hoja de ruta"], obligatorio=False)
    col_patente = buscar_columna(bo, ["Patente de Camión", "PATENTE DE CAMIÓN", "Patente", "Camiones.Patente"], obligatorio=False)
    col_destino = buscar_columna(bo, ["Destino", "DESTINO"], obligatorio=False)

    out = pd.DataFrame(index=bo.index)
    out["Fecha_BO"] = pd.to_datetime(bo[col_fecha], errors="coerce").dt.date
    out["Fecha_key"] = pd.to_datetime(bo[col_fecha], errors="coerce").dt.normalize()
    out["Cliente_BO"] = bo[col_cliente].apply(limpiar_texto)
    out["Contrato_BO"] = bo[col_contrato].apply(normalizar_contrato)
    out["Peso_BO"] = bo[col_peso].apply(normalizar_numero)
    out["Ticket_BO"] = bo[col_ticket].apply(limpiar_texto) if col_ticket else ""
    out["Patente_BO"] = bo[col_patente].apply(limpiar_texto) if col_patente else ""
    out["Destino_BO"] = bo[col_destino].apply(limpiar_texto) if col_destino else ""
    out["Cliente_key"] = out["Cliente_BO"].apply(normalizar_texto)
    out["Contrato_key"] = out["Contrato_BO"].apply(normalizar_contrato)
    return out.dropna(how="all").reset_index(drop=True)


def tomar_primer_match(sub):
    if sub.empty:
        return {
            "Fecha_BO": "",
            "Cliente_BO": "",
            "Contrato_BO": "",
            "Peso_BO": np.nan,
            "Diferencia_kg": np.nan,
            "Ticket_BO": "",
            "Patente_BO": "",
            "Destino_BO": "",
        }
    r = sub.iloc[0]
    return {
        "Fecha_BO": r.get("Fecha_BO", ""),
        "Cliente_BO": r.get("Cliente_BO", ""),
        "Contrato_BO": r.get("Contrato_BO", ""),
        "Peso_BO": r.get("Peso_BO", np.nan),
        "Diferencia_kg": np.nan,
        "Ticket_BO": r.get("Ticket_BO", ""),
        "Patente_BO": r.get("Patente_BO", ""),
        "Destino_BO": r.get("Destino_BO", ""),
    }


def comparar_fila(fila, bo):
    fecha = fila["Fecha_key"]
    cliente = fila["Cliente_key"]
    contrato = fila["Contrato_key"]
    peso = fila["Pesaje_SR"]

    base = {
        "Resultado": "",
        "Motivo": "",
        "Fecha_BO": "",
        "Cliente_BO": "",
        "Contrato_BO": "",
        "Peso_BO": np.nan,
        "Diferencia_kg": np.nan,
        "Ticket_BO": "",
        "Patente_BO": "",
        "Destino_BO": "",
    }

    faltan = []
    if pd.isna(fecha):
        faltan.append("fecha")
    if not cliente:
        faltan.append("cliente")
    if not contrato:
        faltan.append("contrato")
    if pd.isna(peso):
        faltan.append("pesaje")
    if faltan:
        base["Resultado"] = "NO COMPARABLE"
        base["Motivo"] = "Falta " + ", ".join(faltan) + " en Simple Route"
        return pd.Series(base)

    bo_fecha = bo[bo["Fecha_key"].eq(fecha)]
    if bo_fecha.empty:
        base["Resultado"] = "NO ENCONTRADO EN BO"
        base["Motivo"] = "No existe ningún registro BO en esa fecha"
        return pd.Series(base)

    bo_fcc = bo_fecha[(bo_fecha["Cliente_key"].eq(cliente)) & (bo_fecha["Contrato_key"].eq(contrato))].copy()
    exacto = bo_fcc[bo_fcc["Peso_BO"].apply(lambda x: pesos_equivalentes(peso, x))].copy()
    if not exacto.empty:
        match = tomar_primer_match(exacto)
        base.update(match)
        base["Resultado"] = "OK"
        base["Motivo"] = "Existe en BO con mismo cliente, contrato y pesaje"
        base["Diferencia_kg"] = round(float(peso) - float(match["Peso_BO"]), 2)
        return pd.Series(base)

    if not bo_fcc.empty:
        bo_fcc["diff"] = (bo_fcc["Peso_BO"] - peso).abs()
        cercano = bo_fcc.sort_values("diff").head(1)
        match = tomar_primer_match(cercano)
        base.update(match)
        base["Resultado"] = "PESO DISTINTO"
        base["Motivo"] = "Coincide fecha + cliente + contrato, pero no coincide pesaje"
        base["Diferencia_kg"] = round(float(peso) - float(match["Peso_BO"]), 2) if pd.notna(match["Peso_BO"]) else np.nan
        return pd.Series(base)

    bo_cliente_peso = bo_fecha[(bo_fecha["Cliente_key"].eq(cliente)) & (bo_fecha["Peso_BO"].apply(lambda x: pesos_equivalentes(peso, x)))].copy()
    if not bo_cliente_peso.empty:
        match = tomar_primer_match(bo_cliente_peso)
        base.update(match)
        base["Resultado"] = "CONTRATO DISTINTO"
        base["Motivo"] = "Coincide fecha + cliente + pesaje, pero contrato es distinto"
        base["Diferencia_kg"] = round(float(peso) - float(match["Peso_BO"]), 2) if pd.notna(match["Peso_BO"]) else np.nan
        return pd.Series(base)

    bo_contrato_peso = bo_fecha[(bo_fecha["Contrato_key"].eq(contrato)) & (bo_fecha["Peso_BO"].apply(lambda x: pesos_equivalentes(peso, x)))].copy()
    if not bo_contrato_peso.empty:
        match = tomar_primer_match(bo_contrato_peso)
        base.update(match)
        base["Resultado"] = "CLIENTE DISTINTO"
        base["Motivo"] = "Coincide fecha + contrato + pesaje, pero cliente es distinto"
        base["Diferencia_kg"] = round(float(peso) - float(match["Peso_BO"]), 2) if pd.notna(match["Peso_BO"]) else np.nan
        return pd.Series(base)

    bo_cliente = bo_fecha[bo_fecha["Cliente_key"].eq(cliente)].copy()
    if not bo_cliente.empty:
        match = tomar_primer_match(bo_cliente)
        base.update(match)
        base["Resultado"] = "NO ENCONTRADO EN BO"
        base["Motivo"] = "Existe el cliente ese día, pero no coincide contrato ni pesaje"
        return pd.Series(base)

    bo_contrato = bo_fecha[bo_fecha["Contrato_key"].eq(contrato)].copy()
    if not bo_contrato.empty:
        match = tomar_primer_match(bo_contrato)
        base.update(match)
        base["Resultado"] = "NO ENCONTRADO EN BO"
        base["Motivo"] = "Existe el contrato ese día, pero no coincide cliente ni pesaje"
        return pd.Series(base)

    base["Resultado"] = "NO ENCONTRADO EN BO"
    base["Motivo"] = "No existe coincidencia por fecha, cliente, contrato ni pesaje"
    return pd.Series(base)


def comparar(df_sr, bo):
    if df_sr.empty:
        return df_sr.copy()
    diag = df_sr.apply(lambda fila: comparar_fila(fila, bo), axis=1)
    out = pd.concat([df_sr.reset_index(drop=True), diag.reset_index(drop=True)], axis=1)
    columnas_finales = [
        "Fecha_SR",
        "Estado_SR",
        "Contrato_SR",
        "Tipo_SR",
        "Cliente_SR",
        "Pesaje_SR",
        "Resultado",
        "Motivo",
        "Fecha_BO",
        "Cliente_BO",
        "Contrato_BO",
        "Peso_BO",
        "Diferencia_kg",
        "Ticket_BO",
        "Patente_BO",
        "Destino_BO",
        "Título",
        "ID de visita",
        "Tracking ID",
    ]
    columnas_finales = [c for c in columnas_finales if c in out.columns]
    return out[columnas_finales].copy()


def resumen(sr_limpia, df_comparar, comp=None, comparar_todos_los_estados=False):
    data = [
        ["Pregunta de control", "Todo lo que Simple Route dice que fue DF, ¿existe en BO San Bernardo con el mismo cliente, contrato y pesaje?"],
        ["Criterio Simple Route", "DF en columna Título"],
        ["Filtro de estado", "Todos los estados" if comparar_todos_los_estados else "Solo Estado = Completed"],
        ["Total filas Simple Route", len(sr_limpia)],
        ["Total filas DF detectadas", int(sr_limpia["Es_DF"].sum())],
        ["Filas DF comparadas", len(df_comparar)],
        ["DF comparadas sin pesaje", int(df_comparar["Pesaje_SR"].isna().sum()) if not df_comparar.empty else 0],
        ["Títulos DF a revisar", int(((sr_limpia["Es_DF"]) & (sr_limpia["Estado_titulo"] != "OK")).sum())],
    ]
    if comp is not None and not comp.empty:
        for resultado, n in comp["Resultado"].value_counts(dropna=False).items():
            data.append([f"Resultado - {resultado}", int(n)])
    return pd.DataFrame(data, columns=["Indicador", "Valor"])


def estilizar(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border = border
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells[:250]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)
        ws.row_dimensions[1].height = 30

    if "COMPARACION_DF" in wb.sheetnames:
        ws = wb["COMPARACION_DF"]
        resultado_col = None
        for cell in ws[1]:
            if cell.value == "Resultado":
                resultado_col = cell.column
                break
        if resultado_col:
            for row in range(2, ws.max_row + 1):
                val = str(ws.cell(row, resultado_col).value or "").upper()
                fill = ok_fill if val == "OK" else warn_fill if "DISTINTO" in val or "NO COMPARABLE" in val else bad_fill
                ws.cell(row, resultado_col).fill = fill
                ws.cell(row, resultado_col).font = Font(bold=True)

    wb.save(path)
    wb.close()


def controlar(ruta_sr, ruta_bo=None, ruta_salida=None,
              comparar_todos_los_estados=False, mostrar=True):
    """Ejecuta el control y devuelve los tres cuadros más el registro.

    ruta_sr     — archivo exportado de Simple Route
    ruta_bo     — BBDD BO SAN BERNARDO.xlsx. Si falta, el control igual corre
                  y marca todas las filas como "SIN BO"
    ruta_salida — dónde escribir el Excel. Si es None, no se escribe archivo
    """
    p = Registro(mostrar=mostrar)

    path_sr = Path(ruta_sr)
    if not path_sr.exists():
        raise FileNotFoundError(
            f"No existe el archivo de Simple Route: {path_sr}"
        )

    p("  Leyendo Simple Route...")
    sr_limpia, df_comparar, titulos_revisar = preparar_simple_route(
        path_sr, comparar_todos_los_estados
    )
    p(f"    {len(sr_limpia)} filas · {int(sr_limpia['Es_DF'].sum())} marcadas DF "
      f"· {len(df_comparar)} a comparar")

    path_bo = Path(ruta_bo) if ruta_bo else None
    if path_bo is None or not path_bo.exists():
        p("  ⚠ No se encontró la BO de San Bernardo: no hay contra qué comparar.")
        comp = df_comparar.copy()
        comp["Resultado"] = "SIN BO"
        comp["Motivo"] = "No se encontró la BO para comparar"
    else:
        p(f"  Leyendo {path_bo.name}...")
        bo = preparar_bo(path_bo)
        p(f"    {len(bo)} filas")
        p("  Comparando...")
        comp = comparar(df_comparar, bo)

    res = resumen(sr_limpia, df_comparar, comp, comparar_todos_los_estados)

    archivo = None
    if ruta_salida:
        path_salida = Path(ruta_salida)
        with pd.ExcelWriter(path_salida, engine="openpyxl") as writer:
            res.to_excel(writer, sheet_name="RESUMEN", index=False)
            comp.to_excel(writer, sheet_name="COMPARACION_DF", index=False)
            titulos_revisar.to_excel(writer, sheet_name="TITULOS_A_REVISAR", index=False)
        estilizar(path_salida)
        archivo = path_salida.name
        p(f"  ✓ Guardado en: {archivo}")

    conteo = {}
    if not comp.empty and "Resultado" in comp.columns:
        conteo = {str(k): int(v) for k, v in
                  comp["Resultado"].value_counts(dropna=False).items()}
        p("")
        p("  Resultados:")
        for k, v in conteo.items():
            p(f"    {k:<28} {v:>6}")

    return {
        "resumen": res,
        "comparacion": comp,
        "titulos_revisar": titulos_revisar,
        "conteo": conteo,
        "archivo": archivo,
        "log": p.texto(),
    }


def _main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "simple_route",
        nargs="?",
        default="simple_route_rm.xlsx",
        help="Ruta archivo exportado desde Simple Route. Si no se indica, usa ./simple_route_rm.xlsx dentro de la carpeta RM.",
    )
    parser.add_argument("--bo", default=None, help="Ruta BBDD BO SAN BERNARDO.xlsx")
    parser.add_argument("--salida", default=NOMBRE_SALIDA, help="Ruta archivo de salida")
    parser.add_argument("--comparar-todos-los-estados", action="store_true", help="Compara todos los DF, no solo Estado=Completed")
    args = parser.parse_args()

    controlar(
        args.simple_route,
        ruta_bo=args.bo,
        ruta_salida=args.salida,
        comparar_todos_los_estados=args.comparar_todos_los_estados,
        mostrar=True,
    )


if __name__ == "__main__":
    _main()
