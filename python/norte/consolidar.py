"""
consolidar.py — Consolidador de Trazabilidad Zona Norte (versión refactorizada)
Ambipar Group — Proyecto TRZ-APP-001, Etapa 1

Mismo patrón que src/rm/consolidar.py y src/sur/consolidar.py: la lógica de
las 3 fuentes, el cruce con SINADER y el cruce (de dos pasadas) con
Destinatarios son exactamente los mismos que en 02_consolidar_norte.py,
línea por línea.

Diferencias reales respecto de Sur (no son errores, están documentadas en el
script original):
    - PLANTAS_PROPIAS_ACTIVAS SÍ se usa en Norte (a diferencia de RM, donde
      la constante existe pero no participa en ningún cálculo). RECYNOR y
      RECYNOR ARICA son plantas activas; un traslado hacia ellas se descarta
      del lado de quien despacha.
    - Al cargar Destinatarios, Norte NO excluye la Región Metropolitana (Sur
      sí la excluye). Una BO del norte puede despachar a un destinatario
      metropolitano y ese cruce debe funcionar.
    - unir_destinatarios tiene solo DOS pasadas (con región, sin región), no
      las tres de Sur — Norte no tiene la pasada de respaldo "solo destino +
      LER, casos inequívocos".
    - El establecimiento de Arica se normaliza siempre a "RECYNOR ARICA"
      (nombre de Ventanilla Única), aunque en los archivos operativos
      aparezca como "IRAR ARICA".

Se puede usar de dos formas:

    Desde otro programa:
        from consolidar import consolidar, rutas_desde_config
        rutas = rutas_desde_config()
        resultado = consolidar(rutas, modo_prueba=True)

    Desde la terminal, igual que antes:
        python consolidar.py --prueba
        python consolidar.py --reset
"""

import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill


# ══════════════════════════════════════════════════════════════════
# CONSTANTES DE LA ZONA NORTE — reglas de negocio, no rutas
# ══════════════════════════════════════════════════════════════════
FECHA_DESDE_NORTE = "2026-07-01"
HOJA_DESTINO = "TRAZABILIDAD"

REGION_RECYNOR    = "Región de Tarapacá"
REGION_BO_IQUIQUE = "Región de Tarapacá"
REGION_IRAR_ARICA = "Región de Arica y Parinacota"

# El establecimiento de Arica está inscrito en Ventanilla Única como
# "RECYNOR ARICA": nombre canónico que debe salir al consolidado y el que
# debe existir en BBDD_DESTINATARIO. "IRAR ARICA" es el nombre operativo
# interno del mismo lugar; ambos se normalizan a RECYNOR ARICA.
PLANTA_IRAR_ARICA = "RECYNOR ARICA"

HOJA_DESTINATARIOS = "NACIONAL"

# Plantas propias activas en Norte. A diferencia de RM, esta constante SÍ se
# usa: ver excluir_traslados_a_planta_propia().
PLANTAS_PROPIAS_ACTIVAS = {
    "RECYNOR",
    "RECYNOR ARICA",
}

COLUMNAS_FINALES = [
    "Fecha", "Mes", "Cliente", "RUT", "Gestor", "Contrato", "Generador",
    "Transportista", "Rut transportista", "Patente de Camión", "ID", "Origen ID",
    "Peso neto (kg)", "Destino", "Comuna Destino", "RUT DESTINATARIO",
    "TIPO", "CÓDIGOS SINADER", "Movimiento", "Movimiento interempresa",
    "CÓDIGO ESTABLECIMIENTO SINADER", "CÓDIGO DE TRATAMIENTO SINADER",
    "Región", "Destino inferido",
]

# El ticket sale del archivo, así que la clave pasa a usar ID. Se suma
# "Peso neto (kg)" porque una guía puede amparar varios movimientos con
# distinto peso: sin el peso, esas filas legítimas se leerían como
# duplicadas. Medido sobre las bases actuales, esta clave marca MENOS
# duplicados que la anterior en todas las fuentes.
CLAVE_DEDUP = ["Fecha", "Cliente", "ID", "TIPO", "Destino", "Peso neto (kg)"]

MAX_REINTENTOS = 5
ESPERA_REINTENTO = 30

MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


class ArchivoFaltante(Exception):
    """Se lanza cuando falta algún archivo fuente obligatorio."""


# ══════════════════════════════════════════════════════════════════
# REGISTRO DE MENSAJES
# ══════════════════════════════════════════════════════════════════
class Registro:
    """Ver docstring de la clase equivalente en src/rm/consolidar.py."""

    def __init__(self, mostrar=True, ruta_log=None):
        self.lineas = []
        self.mostrar = mostrar
        self.ruta_log = Path(ruta_log) if ruta_log else None

    def __call__(self, msg=""):
        msg = str(msg)
        self.lineas.append(msg)
        if self.mostrar:
            print(msg)
        if self.ruta_log:
            marca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(self.ruta_log, "a", encoding="utf-8") as f:
                f.write(f"{marca}  {msg}\n")

    def texto(self):
        return "\n".join(self.lineas)


# ══════════════════════════════════════════════════════════════════
# CÓMO ENCONTRAR LOS ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def rutas_desde_config(base=None):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas

    c = carpetas(base)
    zona = c["zona_norte"]
    info = c["info_base"]

    rutas = {
        "recynor":        zona / "BBDD RECYNOR.xlsx",
        "bo_iquique":      zona / "BBDD BO IQUIQUE.xlsx",
        "irar_arica":      zona / "BBDD IRAR ARICA.xlsx",
        "homologacion":    info / "HOMOLOGACION.xlsx",
        "destinatarios":   info / "BBDD_DESTINATARIO.xlsx",
        "sinader":         info / "Clasificación_Residuos SINADER.xlsx",
        "transportistas":  info / "Transportistas.xlsx",
        "destino_real":    c["bbdd"] / "BBDD_TRAZABILIDAD_NORTE.xlsx",
    }
    alt_dest = info / "BBDD DESTINATARIO.xlsx"
    if not rutas["destinatarios"].exists() and alt_dest.exists():
        rutas["destinatarios"] = alt_dest

    return rutas


def verificar_rutas(rutas):
    """irar_arica es opcional — el script original la omite si no existe."""
    obligatorias = [
        "recynor", "bo_iquique", "homologacion", "destinatarios",
        "sinader", "transportistas",
    ]
    faltan = [f"{k}: {rutas[k]}" for k in obligatorias if k not in rutas or not Path(rutas[k]).exists()]
    if faltan:
        detalle = "\n".join(f"    · {f}" for f in faltan)
        raise ArchivoFaltante(f"No se encontraron estos archivos:\n{detalle}")


# ══════════════════════════════════════════════════════════════════
# FUNCIONES DE TEXTO — idénticas a 02_consolidar_norte.py
# ══════════════════════════════════════════════════════════════════
def nombre_mes(fecha):
    try:
        return MESES_ES[pd.Timestamp(fecha).month]
    except Exception:
        return ""


def limpiar_texto(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a"}:
        return None
    return re.sub(r"\s+", " ", s).strip()


def normalizar_texto(x):
    s = limpiar_texto(x)
    if not s:
        return ""
    s = s.upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = s.replace(".", "")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalizar_nombre_para_buscar(x):
    return normalizar_texto(x)


def normalizar_region_para_cruce(x):
    """Normaliza regiones para poder cruzar (ej. 'Región de Tarapacá' -> 'TARAPACA')."""
    n = normalizar_texto(x)
    n = n.replace("REGION DE ", "")
    n = n.replace("REGION DEL ", "")
    n = n.replace("REGION ", "")
    n = n.replace("METROPOLITANA DE SANTIAGO", "METROPOLITANA")
    return n.strip()


def es_cliente_resimple(x):
    """Detecta RESIMPLE aunque venga como RE SIMPLE, RE-SIMPLE, RESIMPLE S.A., etc."""
    n = normalizar_texto(x)
    n_sin_espacios = re.sub(r"[^A-Z0-9]", "", n)
    return "RESIMPLE" in n_sin_espacios


def limpiar_numero(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a", "-"}:
        return None
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
        s = s.replace(".", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", ".", "-", "-."}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def normalizar_movimiento(x, default=None):
    s = limpiar_texto(x)
    if not s:
        return default
    n = normalizar_texto(s)
    if "INGRES" in n:
        return "Ingreso"
    if "TRASL" in n:
        return "Traslado"
    if "SALIDA" in n:
        return "Salida"
    return s.strip()


def normalizar_destino(x, planta_si_ecofibras=None):
    s = limpiar_texto(x)
    if not s:
        return None
    n = normalizar_texto(s)
    # Establecimiento de Arica: en Ventanilla Única es "RECYNOR ARICA".
    # "IRAR ARICA" es el nombre operativo interno del mismo lugar. No
    # confundir con la planta RECYNOR de Iquique.
    if "RECYNOR" in n and "ARICA" in n:
        return "RECYNOR ARICA"
    if "RECYNOR" in n:
        return "RECYNOR"
    if "IRAR ARICA" in n or ("IRAR" in n and "ARICA" in n):
        return "RECYNOR ARICA"
    if "I R A R" in n or "IRAR" in n:
        return "IRAR LOS ÁNGELES"
    if "GIRI" in n:
        return "GIRI"
    if "ECOFIBRAS SAN BERNARDO" in n:
        return "ECOFIBRAS SAN BERNARDO"
    if n in {"ECOFIBRAS", "ECOFIBRAS SA", "ECOFIBRAS S A"} and planta_si_ecofibras:
        return planta_si_ecofibras
    return s.strip().upper()


def comuna_por_destino(destino):
    d = normalizar_texto(destino)
    if not d:
        return None
    if "RECYNOR" in d and "ARICA" in d:
        return "ARICA"
    if "RECYNOR" in d:
        return "IQUIQUE"
    if "IRAR ARICA" in d:
        return "ARICA"
    if "IRAR" in d or "LOS ANGELES" in d:
        return "LOS ÁNGELES"
    if "GIRI" in d:
        return "QUILICURA"
    if "ECOFIBRAS SAN BERNARDO" in d:
        return "SAN BERNARDO"
    return None


def buscar_columna(df, opciones, obligatorio=False):
    mapa = {normalizar_texto(c): c for c in df.columns}
    for op in opciones:
        key = normalizar_texto(op)
        if key in mapa:
            return mapa[key]
    if obligatorio:
        raise ValueError(f"No encontré ninguna columna entre: {opciones}. Columnas disponibles: {list(df.columns)}")
    return None


def tomar_columna(df, opciones, default=None, obligatorio=False):
    col = buscar_columna(df, opciones, obligatorio=obligatorio)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


# ══════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def leer_excel(path, p, **kwargs):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            return pd.read_excel(path, **kwargs)
        except PermissionError:
            p(f"  ⚠ Bloqueado: {Path(path).name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def leer_fuente(path, hoja, p, headers_fallback=(0, 7, 8, 9)):
    path = Path(path)
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
            if hoja not in wb.sheetnames:
                wb.close()
                raise ValueError(f"No existe la hoja '{hoja}' en {path.name}. Hojas: {wb.sheetnames}")
            ws = wb[hoja]
            if ws.tables:
                tabla = list(ws.tables.values())[0]
                filas = [[c.value for c in fila] for fila in ws[tabla.ref]]
                headers = filas[0]
                df = pd.DataFrame(filas[1:], columns=headers)
                wb.close()
                df.columns = df.columns.astype(str).str.strip()
                return df.dropna(how="all").reset_index(drop=True)
            wb.close()

            mejor = None
            mejor_score = -1
            claves = [
                "FECHA", "CLIENTE", "RUT", "GENERADOR", "TRANSPORTISTA",
                "DESTINO", "TIPO", "MOVIMIENTO", "PESO", "TICKET", "PATENTE",
            ]
            for header in headers_fallback:
                try:
                    cand = pd.read_excel(path, sheet_name=hoja, header=header)
                    cand.columns = cand.columns.astype(str).str.strip()
                    cand = cand.dropna(how="all").reset_index(drop=True)
                    score = 0
                    cols_norm = [normalizar_texto(c) for c in cand.columns]
                    for clave in claves:
                        if any(clave in c for c in cols_norm):
                            score += 1
                    if score > mejor_score:
                        mejor = cand
                        mejor_score = score
                except Exception:
                    pass
            if mejor is None:
                raise ValueError(f"No se pudo leer {path.name} / hoja {hoja}")
            return mejor
        except PermissionError:
            p(f"  ⚠ Bloqueado: {path.name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def cargar_homologacion(ruta_homologacion, p):
    homolog_clientes = {}
    homolog_generadores = {}

    ruta = Path(ruta_homologacion)
    p("\n── Cargando HOMOLOGACION desde OneDrive ──")
    p(f"  {ruta}")

    if not ruta.exists():
        p("  ⚠ No se encontró HOMOLOGACION.xlsx — se continúa sin correcciones")
        return homolog_clientes, homolog_generadores

    try:
        df_cli = leer_excel(ruta, p, sheet_name="clientes")
        df_cli = df_cli.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_cli.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            rut_correcto = limpiar_texto(row.get("RUT_CORRECTO"))
            if variante and nombre_correcto:
                homolog_clientes[variante] = {"nombre": nombre_correcto, "rut": rut_correcto}
        p(f"  Clientes homologados: {len(homolog_clientes)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja clientes: {e}")

    try:
        df_gen = leer_excel(ruta, p, sheet_name="generadores")
        df_gen = df_gen.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_gen.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            if variante and nombre_correcto:
                homolog_generadores[variante] = nombre_correcto
        p(f"  Generadores homologados: {len(homolog_generadores)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja generadores: {e}")

    return homolog_clientes, homolog_generadores


def aplicar_homologacion_cliente(df, col_nombre, col_rut, homolog_clientes):
    if not homolog_clientes or col_nombre not in df.columns:
        return df
    if col_rut not in df.columns:
        df[col_rut] = None

    def corregir(row):
        variante = normalizar_nombre_para_buscar(row.get(col_nombre))
        if variante in homolog_clientes:
            corr = homolog_clientes[variante]
            nuevo_nombre = corr["nombre"]
            nuevo_rut = corr["rut"] if corr["rut"] else row.get(col_rut)
            return pd.Series([nuevo_nombre, nuevo_rut])
        return pd.Series([row.get(col_nombre), row.get(col_rut)])

    df[[col_nombre, col_rut]] = df.apply(corregir, axis=1)
    return df


def aplicar_homologacion_generador(df, col_generador, homolog_generadores):
    if not homolog_generadores or col_generador not in df.columns:
        return df

    def corregir(x):
        key = normalizar_nombre_para_buscar(x)
        return homolog_generadores.get(key, x)

    df[col_generador] = df[col_generador].apply(corregir)
    return df


def cargar_lookups(rutas, p):
    p("\n── Cargando maestros compartidos desde OneDrive ──")

    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    df_t = df_t[["Transportista", "RUT"]].dropna(subset=["Transportista"])
    df_t["Transportista"] = df_t["Transportista"].astype(str).str.strip()
    df_t["RUT"] = df_t["RUT"].astype(str).str.strip()
    transportistas = dict(zip(df_t["Transportista"], df_t["RUT"]))
    p(f"  Transportistas: {len(transportistas)}")

    df_m = leer_excel(rutas["sinader"], p, sheet_name="Clasificación Residuos", header=1)
    df_m.columns = df_m.columns.astype(str).str.strip()
    df_m = df_m[["TIPO", "CÓDIGOS SINADER"]].dropna(subset=["TIPO"])
    df_m["TIPO"] = df_m["TIPO"].astype(str).str.strip()
    df_m["CÓDIGOS SINADER"] = df_m["CÓDIGOS SINADER"].astype(str).str.strip()
    df_m = df_m.drop_duplicates(subset=["TIPO"])
    materiales = dict(zip(df_m["TIPO"], df_m["CÓDIGOS SINADER"]))
    if "ASIMILABLE A DOMICILIARIO" not in materiales:
        materiales["ASIMILABLE A DOMICILIARIO"] = "20 03 01"
    p(f"  SINADER: {len(materiales)} tipos")

    df_d = leer_excel(rutas["destinatarios"], p, sheet_name=HOJA_DESTINATARIOS, header=1)
    df_d.columns = df_d.columns.astype(str).str.strip()

    col_region_dest = buscar_columna(
        df_d,
        ["REGIÓN DE ESTABLECIMIETNO", "REGIÓN DE ESTABLECIMIENTO", "REGION DE ESTABLECIMIENTO", "REGIÓN", "REGION"],
        obligatorio=True,
    )

    # A diferencia de Sur, aquí NO se excluye la RM: una BO del norte puede
    # despachar a un destinatario metropolitano y ese cruce debe funcionar.
    df_d["_REGION_DEST_NORM"] = df_d[col_region_dest].apply(normalizar_region_para_cruce)
    p(f"  Destinatarios disponibles: {len(df_d)}")

    regiones_norte = {"TARAPACA", "ANTOFAGASTA", "ARICA Y PARINACOTA", "ATACAMA", "COQUIMBO"}
    n_norte = int(df_d["_REGION_DEST_NORM"].isin(regiones_norte).sum())
    if n_norte == 0:
        p("  ⚠ La hoja NACIONAL no tiene ningún destinatario del norte del país.")
        p("     Todas las filas de esta zona van a quedar sin RUT DESTINATARIO")
        p("     ni código de establecimiento SINADER hasta que se carguen.")
    else:
        p(f"  Destinatarios de regiones del norte: {n_norte}")

    cols = [
        "NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", col_region_dest, "CÓDIGOS LER",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER",
    ]
    df_d = df_d[cols].dropna(subset=["NOMBRE DE FANTASÍA"])
    for c in cols:
        df_d[c] = df_d[c].astype(str).str.strip()
    df_d = df_d.rename(columns={col_region_dest: "REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"})

    p(f"  Destinatarios disponibles para cruce LER+Región: {len(df_d)}")

    return transportistas, materiales, df_d


# ══════════════════════════════════════════════════════════════════
# FORMATO DEL ENCABEZADO
# ══════════════════════════════════════════════════════════════════
# Estas columnas se destacan con otro color porque no vienen tal cual de
# la planilla de origen: son derivadas o inferidas por el consolidador.
# Verlas distinto al abrir el archivo evita tratarlas como dato de terreno.
COLUMNAS_DESTACADAS = {
    "Origen ID",
    "Comuna Destino",
    "Movimiento",
    "Movimiento interempresa",
    "Destino inferido",
}

# Paleta Ambipar
COLOR_TEAL = "FF032024"   # fondo del encabezado normal
COLOR_LIMA = "FFCDFF00"   # texto del encabezado normal
COLOR_CREMA = "FFF5F4ED"  # fondo de las columnas destacadas


def formatear_encabezado(ws, columnas):
    """Pinta la fila 1: teal por defecto, crema en las columnas derivadas."""
    for i, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=i)
        if nombre in COLUMNAS_DESTACADAS:
            celda.font = Font(bold=True, color=COLOR_TEAL)
            celda.fill = PatternFill("solid", fgColor=COLOR_CREMA)
        else:
            celda.font = Font(bold=True, color=COLOR_LIMA)
            celda.fill = PatternFill("solid", fgColor=COLOR_TEAL)
    ws.freeze_panes = "A2"

def guardar_excel(df, ruta, p):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            if ruta.exists():
                wb = load_workbook(ruta)
                if HOJA_DESTINO in wb.sheetnames:
                    del wb[HOJA_DESTINO]
                ws = wb.create_sheet(HOJA_DESTINO)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = HOJA_DESTINO
            ws.append(list(df.columns))
            for fila in df.itertuples(index=False):
                ws.append([None if str(v) in ("nan", "NaT", "None") else v for v in fila])

            formatear_encabezado(ws, list(df.columns))
            wb.save(ruta)
            p(f"  ✓ Guardado en: {ruta}")
            return
        except PermissionError:
            p(f"  ⚠ Archivo bloqueado — intento {intento}/{MAX_REINTENTOS}: {ruta.name}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo guardar {ruta}. Ciérralo en Excel.")



# ══════════════════════════════════════════════════════════════════
# ID DE TRAZABILIDAD
# ══════════════════════════════════════════════════════════════════
# El consolidado ya no publica "Ticket de pesaje". En su lugar sale "ID",
# que toma el primer valor disponible en este orden de prioridad:
#     1) N° Guía Cliente   2) N° Guía Ecofibras   3) Ticket de pesaje
# y "N/A" si ninguna de las tres trae dato.
#
# "Origen ID" registra cuál de las tres se usó. Sin esa columna, un número
# en ID puede ser tres documentos distintos y no habría forma de saber cuál,
# que es justo lo que importa al reportar a SINADER.
#
# Cada procesar_*() deja las columnas de trabajo _guia_cliente y
# _guia_ecofibras; construir_id() las consume y luego las elimina.
VALORES_NULOS_ID = {
    "", "-", "--", ".", "0", "S/I", "N/A", "NA", "NONE", "NAN", "NAT",
    "SIN GUIA", "SIN GUÍA", "SIN TICKET", "SIN INFORMACION",
    "SIN INFORMACIÓN", "SIN HR ASOCIADO",
}


def _valor_id(v):
    """Normaliza un identificador. Devuelve None si está vacío o es placeholder."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if s.upper() in VALORES_NULOS_ID:
        return None
    # Excel entrega los correlativos como float: "45954.0" → "45954"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


def construir_id(df, p):
    """Crea ID y Origen ID según la cadena de prioridad y descarta las auxiliares."""
    for col in ("_guia_cliente", "_guia_ecofibras"):
        if col not in df.columns:
            df[col] = None

    vacio = pd.Series([None] * len(df), index=df.index, dtype=object)
    cliente = df["_guia_cliente"].map(_valor_id)
    ecofibras = df["_guia_ecofibras"].map(_valor_id)
    ticket = df["Ticket de pesaje"].map(_valor_id) if "Ticket de pesaje" in df.columns else vacio

    df["ID"] = cliente.fillna(ecofibras).fillna(ticket).fillna("N/A")

    origen = vacio.copy()
    origen = origen.mask(ticket.notna(), "Ticket de pesaje")
    origen = origen.mask(ecofibras.notna(), "Guía Ecofibras")
    origen = origen.mask(cliente.notna(), "Guía cliente")
    df["Origen ID"] = origen.fillna("N/A")

    p("\n── Construcción de ID (Guía cliente → Guía Ecofibras → Ticket) ──")
    total = len(df)
    reparto = df["Origen ID"].value_counts().to_dict()
    for etiqueta in ("Guía cliente", "Guía Ecofibras", "Ticket de pesaje", "N/A"):
        n = int(reparto.get(etiqueta, 0))
        pct = f"{100 * n / total:.1f}%" if total else "0.0%"
        p(f"  • {etiqueta:<18} {n:>6} filas ({pct})")
    n_na = int((df["ID"] == "N/A").sum())
    if n_na:
        p(f"  ⚠ {n_na} filas sin identificador en ninguna de las tres columnas. REVISAR.")

    return df.drop(columns=["_guia_cliente", "_guia_ecofibras"])


# ══════════════════════════════════════════════════════════════════
# FUENTES
# ══════════════════════════════════════════════════════════════════
def base_final(n):
    return pd.DataFrame(index=range(n), columns=[
        "Fecha", "Mes", "Cliente", "RUT", "Gestor", "Contrato", "Generador",
        "Transportista", "Rut transportista", "Patente de Camión", "Ticket de pesaje",
        "_guia_cliente", "_guia_ecofibras",
        "Peso neto", "Unidad", "Destino", "Comuna Destino", "TIPO", "Movimiento",
        "Movimiento interempresa", "Región", "Destino inferido",
    ])


def filtrar_cliente_resimple(df, col_cliente, nombre_fuente, p):
    if col_cliente not in df.columns:
        p(f"  ⚠ No se encontró columna cliente para filtrar RESIMPLE en {nombre_fuente}")
        return df
    mask_resimple = df[col_cliente].apply(es_cliente_resimple)
    n = int(mask_resimple.sum())
    if n > 0:
        p(f"  • Filtro Cliente contiene RESIMPLE: excluidas {n} filas en {nombre_fuente}")
    return df.loc[~mask_resimple].copy()


def excluir_traslados_a_planta_propia(out, nombre_fuente, p):
    if "Destino" not in out.columns or out.empty:
        return out
    plantas_norm = {normalizar_texto(pl) for pl in PLANTAS_PROPIAS_ACTIVAS}
    dest_norm = out["Destino"].apply(normalizar_texto)
    es_traslado = out["Movimiento"].astype(str).str.strip().str.lower().ne("ingreso")
    mask = dest_norm.isin(plantas_norm) & es_traslado
    n = int(mask.sum())
    if n > 0:
        detalle = out.loc[mask].groupby("Destino").size().to_dict()
        p(f"  • Traslados a planta propia descartados en {nombre_fuente}: {n}")
        for destino, cnt in detalle.items():
            p(f"     → {destino}: {cnt} viajes (los registra la planta como Ingreso)")
        p(f"     Si la planta no los registró, esos {n} viajes NO quedan trazados.")
    return out.loc[~mask].copy()


def destino_recynor(row):
    """Recynor opera como planta y como base operacional.

    La columna MOVIMIENTO manda: INGRESO → destino = RECYNOR;
    RETIRO → destino = columna DESTINO (equivale a un traslado de BO).
    """
    if row.get("Movimiento_final") == "Ingreso":
        return "RECYNOR"
    return normalizar_destino(row.get("DESTINO_COL"))


def procesar_recynor(rutas, transportistas, homolog_c, homolog_g, p):
    """Recynor — planta de reciclaje Y base operacional. Iquique, Tarapacá.

    Nota: ECOFIBRAS S.A. aparece como cliente frecuente. Por decisión del
    equipo entra como cliente normal, SIN marca de movimiento interempresa.
    """
    p("\n── Recynor — planta + BO (hoja INGRESOS) ──")

    df = leer_fuente(rutas["recynor"], "INGRESOS", p, headers_fallback=(0, 8, 9, 7))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["FECHA", "Fecha"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_NORTE)].copy()
    p(f"  • Filtro Norte (>= {FECHA_DESDE_NORTE}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    col_cli = buscar_columna(df, ["CLIENTE", "Cliente"], obligatorio=False)
    if col_cli:
        df = filtrar_cliente_resimple(df, col_cli, "Recynor", p)

    out = base_final(len(df))

    mov_col     = tomar_columna(df, ["MOVIMIENTO", "Movimiento"])
    destino_col = tomar_columna(df, ["DESTINO", "Destino", "DESTINO FINAL"])

    tmp = pd.DataFrame(index=df.index)
    # RETIRO no cae en ninguna regla de normalizar_movimiento, así que llega
    # como "RETIRO"; se traduce explícitamente a Traslado.
    tmp["Movimiento_final"] = mov_col.apply(
        lambda x: "Traslado" if "RETIRO" in normalizar_texto(x) else normalizar_movimiento(x, default="Ingreso")
    )
    tmp["DESTINO_COL"] = destino_col

    p(f"  • Distribución Movimiento: {tmp['Movimiento_final'].value_counts(dropna=False).to_dict()}")

    col_dest = buscar_columna(df, ["DESTINO", "Destino"])
    if col_dest is None:
        p("  ⚠ No existe columna DESTINO en Recynor: los RETIRO quedarán sin destino.")
    else:
        n_retiro_sin_destino = int(
            (tmp["Movimiento_final"].eq("Traslado") & tmp["DESTINO_COL"].apply(limpiar_texto).isna()).sum()
        )
        if n_retiro_sin_destino:
            p(f"  ⚠ RETIRO sin destino informado: {n_retiro_sin_destino} filas")
            p("     Esas filas no van a cruzar con BBDD DESTINATARIO. Pedir que se complete la columna DESTINO.")

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["CLIENTE", "Cliente"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["RUT CLIENTE", "RUT"]).apply(limpiar_texto).values
    out["Gestor"]            = tomar_columna(df, ["TRANSPORTISTA", "GESTOR"]).apply(limpiar_texto).values
    out["Contrato"]          = tomar_columna(df, ["ORDEN DE SERVICIO", "CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]         = tomar_columna(df, ["GENERADOR", "Generador"]).apply(limpiar_texto).values
    out["Transportista"]     = tomar_columna(df, ["TRANSPORTISTA", "Transportista"]).apply(limpiar_texto).values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Patente de Camión"] = tomar_columna(df, ["PATENTE CAMIÓN", "PATENTE", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["TICKET DE PESAJE", "Ticket de pesaje"]).apply(limpiar_texto).values

    # Recynor no registra guía: el ID sale del ticket de pesaje.
    out["_guia_cliente"]   = None
    out["_guia_ecofibras"] = None
    out["Peso neto"]         = tomar_columna(df, ["PESO NETO KG", "PESO NETO", "Peso neto"]).apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Movimiento"]        = tmp["Movimiento_final"].values
    out["Destino"]           = tmp.apply(destino_recynor, axis=1).values
    out["Comuna Destino"]    = out["Destino"].apply(comuna_por_destino)
    out["TIPO"]              = tomar_columna(df, ["TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_RECYNOR
    out["Destino inferido"]  = "No"

    # Recynor también actúa como BO: si traslada a otra planta propia activa,
    # ese viaje lo registra la planta receptora.
    out = excluir_traslados_a_planta_propia(out, "Recynor", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    sin_rut = out[out["Rut transportista"].isna()]["Transportista"].dropna().unique()
    if len(sin_rut) > 0:
        p(f"  ⚠ Transportistas sin RUT en maestro: {len(sin_rut)}")
        for g in sin_rut[:5]:
            p(f"     → '{g}'")

    p(f"  ✓ Recynor final: {len(out)} filas")
    return out


def procesar_bo_iquique(rutas, homolog_c, homolog_g, p):
    """BO Iquique — base operacional, Región de Tarapacá. Siempre Traslado."""
    p("\n── BO Iquique — SIEMPRE Traslado (hoja BBDD) ──")

    df = leer_fuente(rutas["bo_iquique"], "BBDD", p, headers_fallback=(0, 7, 8, 9))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["Fecha", "FECHA"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_NORTE)].copy()
    p(f"  • Filtro Norte (>= {FECHA_DESDE_NORTE}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    col_cli = buscar_columna(df, ["Cliente", "CLIENTE"], obligatorio=False)
    if col_cli:
        df = filtrar_cliente_resimple(df, col_cli, "BO Iquique", p)

    out = base_final(len(df))

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["Cliente", "CLIENTE"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["RUT", "RUT CLIENTE"]).apply(limpiar_texto).values
    out["Gestor"]            = "AMBIPAR ENVIRONMENT CHILE"
    out["Contrato"]          = tomar_columna(df, ["Contrato", "N° CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]         = tomar_columna(df, ["Generador", "GENERADOR"]).apply(limpiar_texto).values
    out["Transportista"]     = "AMBIPAR ENVIRONMENT CHILE"
    out["Rut transportista"] = "96824110-9"
    out["Patente de Camión"] = tomar_columna(df, ["Patente de Camión", "PATENTE DE CAMIÓN", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["Ticket de pesaje", "TICKET DE PESAJE"]).apply(limpiar_texto).values

    # BO Iquique no registra guía: el ID sale del ticket de pesaje.
    out["_guia_cliente"]   = None
    out["_guia_ecofibras"] = None
    out["Peso neto"]         = tomar_columna(df, ["Peso neto [KG]", "Peso neto", "PESO NETO KG", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]            = tomar_columna(df, ["Unidad", "UNIDAD"], default="kg").apply(limpiar_texto).fillna("kg").values
    out["Movimiento"]        = "Traslado"
    out["Destino"]           = tomar_columna(df, ["Destino", "DESTINO"]).apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]    = tomar_columna(df, ["Comuna destino", "Comuna Destino", "COMUNA DESTINO"], default=None).apply(limpiar_texto).values
    out["Comuna Destino"]    = out["Comuna Destino"].fillna(out["Destino"].apply(comuna_por_destino))
    out["TIPO"]              = tomar_columna(df, ["TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_BO_IQUIQUE
    out["Destino inferido"]  = "No"

    out = excluir_traslados_a_planta_propia(out, "BO Iquique", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ BO Iquique final: {len(out)} filas")
    return out


def procesar_irar_arica(rutas, transportistas, homolog_c, homolog_g, p):
    """IRAR Arica — Región de Arica y Parinacota. Planta + base operativa.

    Ingreso → destino = RECYNOR ARICA (funciona como planta).
    Traslado → destino = columna DESTINO (funciona como BO).
    SÍ está en PLANTAS_PROPIAS_ACTIVAS.
    """
    p("\n── IRAR Arica — planta + BO (hoja INGRESOS) ──")

    if rutas.get("irar_arica") is None or not Path(rutas["irar_arica"]).exists():
        p("  ⚠ Archivo BBDD IRAR ARICA.xlsx no encontrado — fuente omitida")
        return base_final(0)

    df = leer_fuente(rutas["irar_arica"], "INGRESOS", p, headers_fallback=(0, 7, 8, 9))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["Fecha", "FECHA"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_NORTE)].copy()
    p(f"  • Filtro Norte (>= {FECHA_DESDE_NORTE}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    col_cli = buscar_columna(df, ["Cliente", "CLIENTE"], obligatorio=False)
    if col_cli:
        df = filtrar_cliente_resimple(df, col_cli, "IRAR Arica", p)

    mov_col     = tomar_columna(df, ["Movimiento", "MOVIMIENTO"])
    destino_col = tomar_columna(df, ["DESTINO", "Destino"])

    dist_mov = mov_col.apply(lambda x: normalizar_movimiento(x, default=None)).value_counts(dropna=False).to_dict()
    p(f"  • Distribución Movimiento en archivo: {dist_mov}")

    mov_norm      = mov_col.apply(lambda x: normalizar_movimiento(x, default="Ingreso"))
    destinos_norm = destino_col.apply(lambda x: normalizar_destino(x))
    p(f"  • Destinos declarados: {destinos_norm.value_counts(dropna=False).to_dict()}")

    # Regla: Ingreso → el material se queda en la planta, destino = IRAR ARICA
    #        Traslado → destino = columna DESTINO
    destino_final = [
        PLANTA_IRAR_ARICA if mv == "Ingreso" else dc
        for mv, dc in zip(mov_norm, destinos_norm)
    ]

    incoherentes = [
        dc for mv, dc in zip(mov_norm, destinos_norm)
        if mv == "Ingreso" and dc is not None and dc != PLANTA_IRAR_ARICA
    ]
    if incoherentes:
        det = pd.Series(incoherentes).value_counts().to_dict()
        p(f"  ⚠ {len(incoherentes)} filas con Movimiento = Ingreso pero DESTINO a un tercero: {det}")
        p(f"     Se registran como Ingreso a RECYNOR ARICA. Si el material efectivamente SALIÓ,")
        p(f"     esas filas deben marcarse como Traslado en el archivo fuente. REVISAR.")

    n_traslado_sin_dest = sum(1 for mv, dc in zip(mov_norm, destinos_norm) if mv != "Ingreso" and dc is None)
    if n_traslado_sin_dest:
        p(f"  ⚠ {n_traslado_sin_dest} traslados sin DESTINO: quedan sin destino trazado. REVISAR.")

    tk = tomar_columna(df, ["Ticket pesaje", "Ticket de pesaje", "TICKET DE PESAJE"], default=None)
    n_sin_ticket = int(tk.isna().sum()) if tk is not None else len(df)
    if n_sin_ticket:
        p(f"  ⚠ {n_sin_ticket} filas sin Ticket de pesaje. Como IRAR Arica está en "
          f"PLANTAS_PROPIAS_ACTIVAS, los traslados de otras BO hacia acá se descartan "
          f"asumiendo que esta planta los registra. Sin ticket no hay respaldo. REVISAR.")

    out = base_final(len(df))

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["Cliente", "CLIENTE"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["Rut Cliente", "RUT CLIENTE", "RUT"]).apply(limpiar_texto).values
    out["Gestor"]            = "AMBIPAR ENVIRONMENT CHILE"
    out["Contrato"]          = tomar_columna(df, ["Contrato", "N° CONTRATO"], default=None).apply(limpiar_texto).values
    out["Generador"]         = tomar_columna(df, ["Generador", "GENERADOR"]).apply(limpiar_texto).values
    out["Transportista"]     = tomar_columna(df, ["Transportista", "TRANSPORTISTA"]).apply(limpiar_texto).values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Patente de Camión"] = tomar_columna(df, ["Patente", "PATENTE", "Patente de Camión"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["Ticket pesaje", "Ticket de pesaje", "TICKET DE PESAJE"]).apply(limpiar_texto).values

    # IRAR Arica sí prioriza la guía del cliente. Hoy la columna existe pero
    # está vacía; cuando operaciones la llene, el ID la toma sin tocar código.
    out["_guia_cliente"]   = tomar_columna(df, ["N° Guía Cliente", "N° GUÍA CLIENTE", "N° Guía"]).apply(limpiar_texto).values
    out["_guia_ecofibras"] = None
    out["Peso neto"]         = tomar_columna(df, ["Peso neto [kg]", "PESO NETO KG", "Peso neto"]).apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Destino"]           = destino_final
    out["Comuna Destino"]    = out["Destino"].apply(comuna_por_destino)
    out["TIPO"]              = tomar_columna(df, ["Tipo Residuo", "TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento"]        = mov_norm.values
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_IRAR_ARICA
    out["Destino inferido"]  = "No"

    out = excluir_traslados_a_planta_propia(out, "IRAR Arica", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ IRAR Arica final: {len(out)} filas")
    return out


# ══════════════════════════════════════════════════════════════════
# JOINS
# ══════════════════════════════════════════════════════════════════
def unir_sinader(df, materiales, p):
    p("\n── Join con SINADER (TIPO → CÓDIGOS SINADER) ──")

    df["CÓDIGOS SINADER"] = df["TIPO"].astype(str).str.strip().map(materiales)

    sin = df[df["CÓDIGOS SINADER"].isna() & df["TIPO"].notna()][["__origen__", "TIPO"]]

    if len(sin) > 0:
        p(f"  ⚠ {len(sin)} filas sin código SINADER")
        for _, r in sin.drop_duplicates().head(20).iterrows():
            p(f"     [{r['__origen__']}] {r['TIPO']}")
    else:
        p("  ✓ Todos los TIPO tienen código SINADER")

    return df


def unir_destinatarios(df, df_dest, p):
    """Cruce de dos pasadas (con región / sin región). Norte no tiene la
    tercera pasada de respaldo por destino+LER que sí tiene Sur."""
    p("\n── Join con DESTINATARIOS (Destino + Comuna + LER + Región) ──")

    df = df.copy()
    dj = df_dest.copy()

    if "Comuna Destino" in df.columns:
        lookup_comuna = (
            dj[["NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", "REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"]]
            .drop_duplicates()
            .copy()
        )
        lookup_comuna["_k_dest"]   = lookup_comuna["NOMBRE DE FANTASÍA"].apply(normalizar_texto)
        lookup_comuna["_k_region"] = lookup_comuna["REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"].apply(normalizar_region_para_cruce)

        mask = df["Comuna Destino"].isna() & df["Destino"].notna()
        if mask.any():
            tmp_base = df.loc[mask, ["Destino", "Región"]].copy()
            tmp_base["_k_dest"]   = tmp_base["Destino"].apply(normalizar_texto)
            tmp_base["_k_region"] = tmp_base["Región"].apply(normalizar_region_para_cruce)

            tmp = tmp_base.merge(
                lookup_comuna[["_k_dest", "_k_region", "COMUNA DE ESTABLECIMIENTO"]],
                on=["_k_dest", "_k_region"],
                how="left",
            )
            df.loc[mask, "Comuna Destino"] = tmp["COMUNA DE ESTABLECIMIENTO"].values

    df["_k_dest"]   = df["Destino"].apply(normalizar_texto)
    df["_k_comuna"] = df["Comuna Destino"].apply(normalizar_texto)
    df["_k_ler"]    = df["CÓDIGOS SINADER"].apply(normalizar_texto)
    df["_k_region"] = df["Región"].apply(normalizar_region_para_cruce)

    dj["_k_dest"]   = dj["NOMBRE DE FANTASÍA"].apply(normalizar_texto)
    dj["_k_comuna"] = dj["COMUNA DE ESTABLECIMIENTO"].apply(normalizar_texto)
    dj["_k_ler"]    = dj["CÓDIGOS LER"].apply(normalizar_texto)
    dj["_k_region"] = dj["REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"].apply(normalizar_region_para_cruce)

    cols_merge = [
        "_k_dest", "_k_comuna", "_k_ler", "_k_region",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER",
    ]

    dj_merge = dj[cols_merge].drop_duplicates(subset=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], keep="first")

    n_antes = len(df)

    # Pasada 1: cruce con región
    df = df.merge(dj_merge, on=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], how="left")
    df = df.sort_values("__idx__").drop_duplicates(subset=["__idx__"], keep="first")

    n_match_p1 = df["RUT DESTINATARIO"].notna().sum()
    p(f"  Pasada 1 (con región): {n_match_p1}/{n_antes} matchearon")

    # Pasada 2: sin región
    sin_match = df["RUT DESTINATARIO"].isna()
    if sin_match.any():
        dj_sin_region = dj[
            ["_k_dest", "_k_comuna", "_k_ler", "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER"]
        ].drop_duplicates(subset=["_k_dest", "_k_comuna", "_k_ler"], keep="first")

        tmp_sin = df.loc[sin_match, ["__idx__", "_k_dest", "_k_comuna", "_k_ler"]].merge(
            dj_sin_region, on=["_k_dest", "_k_comuna", "_k_ler"], how="left"
        )

        matched_p2 = tmp_sin["RUT DESTINATARIO"].notna()
        n_match_p2 = matched_p2.sum()
        p(f"  Pasada 2 (sin región): {n_match_p2} adicionales matchearon")

        if n_match_p2 > 0:
            idx_p2 = tmp_sin.loc[matched_p2, "__idx__"].values
            for col in ["RUT DESTINATARIO", "CÓDIGO ESTABLECIMIENTO SINADER", "CÓDIGO DE TRATAMIENTO SINADER"]:
                df.loc[df["__idx__"].isin(idx_p2), col] = (
                    tmp_sin.loc[matched_p2].set_index("__idx__")[col].reindex(idx_p2).values
                )

    df = df.drop(columns=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], errors="ignore")

    p(f"  {'✓' if len(df) == n_antes else '⚠'} JOIN TOTAL: {n_antes} → {len(df)} filas")

    sin = df[df["RUT DESTINATARIO"].isna()][
        ["__origen__", "Destino", "Comuna Destino", "Región", "CÓDIGOS SINADER"]
    ].copy()

    if len(sin) > 0:
        agrupado = (
            sin.groupby(["Destino", "Comuna Destino", "Región", "CÓDIGOS SINADER"], dropna=False)["__origen__"]
            .agg(lambda x: ", ".join(sorted(set(map(str, x)))))
            .reset_index()
            .rename(columns={"__origen__": "ORIGENES"})
        )

        p(f"  ⚠ {len(agrupado)} combinaciones sin match en Destinatarios:")
        for _, r in agrupado.head(30).iterrows():
            p(
                f"     [{r['ORIGENES']}] {r['Destino']} | {r['Comuna Destino']} | "
                f"{r['Región']} | {r['CÓDIGOS SINADER']}"
            )
    else:
        p("  ✓ Todos los destinos encontraron match en Destinatarios")

    return df


def limpiar_filas_invalidas(df, p):
    claves = ["Fecha", "Cliente", "Destino", "TIPO"]
    temp = df[claves].copy()

    for c in claves:
        temp[c] = temp[c].astype(str).str.strip().str.lower()

    vacia = temp.isin(["", "nan", "none", "nat"]).all(axis=1)
    n = int(vacia.sum())

    if n > 0:
        p(f"  ⚠ Filas inválidas eliminadas: {n}")

    return df.loc[~vacia].copy()


# ══════════════════════════════════════════════════════════════════
# PIPELINE
# ══════════════════════════════════════════════════════════════════
def consolidar(rutas, ruta_prueba=None, ruta_log=None, modo_reset=False,
               modo_prueba=False, mostrar=True):
    """Ejecuta la consolidación completa de Norte.

    Ver docstring de la versión RM para el detalle de parámetros y de lo que
    devuelve — el patrón es idéntico.
    """
    p = Registro(mostrar=mostrar, ruta_log=ruta_log)

    verificar_rutas(rutas)
    if modo_prueba and not ruta_prueba:
        raise ValueError("modo_prueba=True requiere indicar ruta_prueba.")

    p("\n" + "═" * 72)
    p("  INICIO CONSOLIDACIÓN ZONA NORTE" + (" — MODO PRUEBA" if modo_prueba else ""))
    p(f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    p(f"  Procesando registros desde: {FECHA_DESDE_NORTE}")
    p("═" * 72)

    homolog_c, homolog_g = cargar_homologacion(rutas["homologacion"], p)
    transportistas, materiales, df_dest = cargar_lookups(rutas, p)

    df_recynor    = procesar_recynor(rutas, transportistas, homolog_c, homolog_g, p)
    df_bo_iquique = procesar_bo_iquique(rutas, homolog_c, homolog_g, p)
    df_irar_arica = procesar_irar_arica(rutas, transportistas, homolog_c, homolog_g, p)

    fuentes = {
        "RECYNOR":    df_recynor,
        "BO_IQUIQUE": df_bo_iquique,
        "IRAR_ARICA": df_irar_arica,
    }

    p("\n── Combinando fuentes Zona Norte ──")

    suma_fuentes = 0
    lista = []

    for nombre, dfi in fuentes.items():
        dfi = dfi.copy()
        dfi["__origen__"] = nombre
        lista.append(dfi)
        p(f"  {nombre:<20} {len(dfi):>6} filas")
        suma_fuentes += len(dfi)

    p(f"  {'─' * 30}")
    p(f"  Suma esperada:       {suma_fuentes:>6} filas")

    df = pd.concat(lista, ignore_index=True)
    df["__idx__"] = df.index

    if len(df) == suma_fuentes:
        p(f"  ✓ Combinado: {len(df)} filas (0 pérdidas)")
    else:
        p(f"  ⚠ Diferencia en combinación: esperadas {suma_fuentes}, obtenidas {len(df)}")

    df["Destino inferido"] = "No"

    if "Movimiento interempresa" not in df.columns:
        df["Movimiento interempresa"] = "No"
    df["Movimiento interempresa"] = df["Movimiento interempresa"].fillna("No")

    df = unir_sinader(df, materiales, p)
    df = unir_destinatarios(df, df_dest, p)

    df = df.drop(columns=["Unidad", "__origen__", "__idx__"], errors="ignore")
    df = df.rename(columns={"Peso neto": "Peso neto (kg)"})

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Mes"]   = df["Fecha"].apply(nombre_mes)

    df = limpiar_filas_invalidas(df, p)

    df = construir_id(df, p)

    for col in COLUMNAS_FINALES:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMNAS_FINALES]

    p(f"\n── DataFrame final NORTE: {len(df)} filas × {len(df.columns)} columnas ──")

    if modo_prueba:
        ruta_prueba = Path(ruta_prueba)
        # Sello de fecha/hora para no pisar corridas anteriores (igual que RM y Sur).
        sello = datetime.now().strftime("%Y-%m-%d_%H%M")
        ruta_prueba = ruta_prueba.with_name(f"{ruta_prueba.stem}_{sello}{ruta_prueba.suffix}")
        p(f"\n── Modo PRUEBA — guardando {ruta_prueba.name} ──")
        guardar_excel(df, ruta_prueba, p)
        df_final = df
    else:
        ruta_destino = Path(rutas["destino_real"])

        if modo_reset:
            p("\n── Modo RESET — reemplazando BBDD_TRAZABILIDAD_NORTE.xlsx ──")
            df_final = df.copy()
        else:
            p("\n── Modo ACUMULATIVO ──")

            if ruta_destino.exists():
                df_existente = leer_excel(ruta_destino, p, sheet_name=HOJA_DESTINO)
                p(f"  Trazabilidad NORTE existente: {len(df_existente)} filas")
            else:
                df_existente = pd.DataFrame(columns=COLUMNAS_FINALES)
                p("  Archivo NORTE final no existe — se creará nuevo")

            if not df_existente.empty:
                temp_exist = df_existente.copy()
                temp_new   = df.copy()

                for col in CLAVE_DEDUP:
                    temp_exist[col] = temp_exist[col].astype(str).str.strip() if col in temp_exist.columns else ""
                    temp_new[col]   = temp_new[col].astype(str).str.strip()   if col in temp_new.columns   else ""

                llave_exist = temp_exist[CLAVE_DEDUP].apply(tuple, axis=1)
                llave_nueva = temp_new[CLAVE_DEDUP].apply(tuple, axis=1)

                df_nuevas = df[~llave_nueva.isin(llave_exist)].copy()
            else:
                df_nuevas = df.copy()

            p(f"  Filas nuevas a agregar: {len(df_nuevas)}")
            p(f"  Duplicadas omitidas:    {len(df) - len(df_nuevas)}")

            df_final = pd.concat([df_existente, df_nuevas], ignore_index=True)

        for col in COLUMNAS_FINALES:
            if col not in df_final.columns:
                df_final[col] = None
        df_final = df_final[COLUMNAS_FINALES]

        p(f"  Total final NORTE: {len(df_final)} filas")
        guardar_excel(df_final, ruta_destino, p)

    p("\n" + "═" * 72)
    p("  CONSOLIDACIÓN NORTE COMPLETADA ✓")
    p("═" * 72 + "\n")

    return {
        "consolidado": df_final,
        "filas": len(df_final),
        "columnas": len(df_final.columns),
        "log": p.texto(),
    }


# ══════════════════════════════════════════════════════════════════
# USO DESDE LA TERMINAL
# ══════════════════════════════════════════════════════════════════
def _main():
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas, ConfiguracionFaltante

    modo_reset = "--reset" in sys.argv
    modo_prueba = "--prueba" in sys.argv

    try:
        rutas = rutas_desde_config()
        c = carpetas()
    except ConfiguracionFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    ruta_prueba = c["norte"] / "PRUEBA_TRAZABILIDAD_NORTE.xlsx"
    ruta_log = c["norte"] / "log_consolidacion_norte.txt"

    try:
        consolidar(
            rutas,
            ruta_prueba=ruta_prueba,
            ruta_log=ruta_log,
            modo_reset=modo_reset,
            modo_prueba=modo_prueba,
        )
    except (ArchivoFaltante, PermissionError, ValueError) as e:
        print(f"\n✗ ERROR CRÍTICO: {e}\n")
        sys.exit(2)


if __name__ == "__main__":
    _main()
