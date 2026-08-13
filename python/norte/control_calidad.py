"""
control_calidad.py — Control de calidad Zona Norte (versión refactorizada)
Ambipar Group — Proyecto TRZ-APP-001, Etapa 1

Mismo patrón que src/sur/control_calidad.py: la lógica es la misma que en
01_control_calidad_norte.py, línea por línea. Cambia solo cómo se reciben
las rutas (parámetro, no rutas fijas de OneDrive) y que la función principal
devuelve los resultados además de escribir el Excel.

Roles de cada fuente Zona Norte (documentado en el script original):
    RECYNOR     planta + BO → INGRESO / RETIRO (RETIRO = Traslado)
    IRAR_ARICA  planta + BO → Ingreso / Traslado
    BO_IQUIQUE  solo BO     → siempre Traslado

NOTA — Hallazgo 6 (11-08-2026): el script original imprime "Control de
calidad SUR generado" en vez de "NORTE" al final. Es texto copiado del
script de Sur, no afecta ningún cálculo. Se preserva tal cual, igual que en
revisar_consolidado.py de esta misma zona.

Se puede usar de dos formas:

    Desde otro programa:
        from control_calidad import controlar, rutas_desde_carpeta
        rutas = rutas_desde_carpeta("C:/carpeta/con/los/excel")
        resultado = controlar(rutas)

    Desde la terminal, igual que antes:
        python control_calidad.py
"""

import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook


# ══════════════════════════════════════════════════════════════════
# CONSTANTES DE LA ZONA NORTE — reglas de negocio, no rutas
# ══════════════════════════════════════════════════════════════════
FECHA_DESDE_NORTE = "2026-07-01"
HOJA_DESTINO = "TRAZABILIDAD"

MAX_REINTENTOS = 5
ESPERA_REINTENTO = 30

NOMBRES_ARCHIVO = {
    "recynor":        ["BBDD RECYNOR.xlsx"],
    "bo_iquique":      ["BBDD BO IQUIQUE.xlsx"],
    "irar_arica":      ["BBDD IRAR ARICA.xlsx"],
    "homologacion":    ["HOMOLOGACION.xlsx"],
    "destinatarios":   ["BBDD_DESTINATARIO.xlsx", "BBDD DESTINATARIO.xlsx"],
    "sinader":         ["Clasificación_Residuos SINADER.xlsx", "Clasificacion_Residuos SINADER.xlsx"],
    "transportistas":  ["Transportistas.xlsx"],
}

# irar_arica es opcional: la zona Norte todavía no la tiene activa
# completamente (ver documento TRZ-APP-001, sección 3.3).
OPCIONALES = {"irar_arica"}


class ArchivoFaltante(Exception):
    """Se lanza cuando falta algún archivo fuente obligatorio."""


# ══════════════════════════════════════════════════════════════════
# REGISTRO DE MENSAJES
# ══════════════════════════════════════════════════════════════════
class Registro:
    def __init__(self, mostrar=True):
        self.lineas = []
        self.mostrar = mostrar

    def __call__(self, msg=""):
        self.lineas.append(str(msg))
        if self.mostrar:
            print(msg)

    def texto(self):
        return "\n".join(self.lineas)


# ══════════════════════════════════════════════════════════════════
# CÓMO ENCONTRAR LOS ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def _clave(nombre):
    n = unicodedata.normalize("NFKD", str(nombre))
    n = n.encode("ascii", "ignore").decode("ascii")
    n = n.upper()
    for c in " ._-()":
        n = n.replace(c, "")
    return n


def rutas_desde_config(base=None):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas

    c = carpetas(base)
    zona = c["zona_norte"]
    info = c["info_base"]

    rutas = {
        "recynor":        zona / "BBDD RECYNOR.xlsx",
        "bo_iquique":      zona / "BBDD BO IQUIQUE.xlsx",
        "irar_arica":      zona / "BBDD IRAR ARICA.xlsx",
        "homologacion":    info / "HOMOLOGACION.xlsx",
        "destinatarios":   info / "BBDD_DESTINATARIO.xlsx",
        "sinader":         info / "Clasificación_Residuos SINADER.xlsx",
        "transportistas":  info / "Transportistas.xlsx",
    }
    alt_dest = info / "BBDD DESTINATARIO.xlsx"
    if not rutas["destinatarios"].exists() and alt_dest.exists():
        rutas["destinatarios"] = alt_dest

    return rutas


def rutas_desde_carpeta(carpeta):
    """Arma las rutas buscando todos los archivos dentro de una sola carpeta.
    irar_arica es opcional; el resto es obligatorio."""
    carpeta = Path(carpeta)
    if not carpeta.exists():
        raise ArchivoFaltante(f"La carpeta no existe:\n  {carpeta}")

    presentes = {_clave(f.name): f for f in carpeta.iterdir() if f.is_file()}

    rutas = {}
    faltantes = []
    for interno, posibles in NOMBRES_ARCHIVO.items():
        encontrado = None
        for nombre in posibles:
            encontrado = presentes.get(_clave(nombre))
            if encontrado:
                break
        if encontrado:
            rutas[interno] = encontrado
        elif interno not in OPCIONALES:
            faltantes.append(posibles[0])

    if faltantes:
        hay = "\n".join(f"    · {f.name}" for f in sorted(presentes.values()))
        falta = "\n".join(f"    · {f}" for f in faltantes)
        raise ArchivoFaltante(
            f"Faltan archivos en la carpeta:\n{falta}\n\n"
            f"  Archivos encontrados en {carpeta.name}:\n{hay if hay else '    (ninguno)'}"
        )

    return rutas


def verificar_rutas(rutas):
    obligatorias = [k for k in NOMBRES_ARCHIVO if k not in OPCIONALES]
    faltan = [f"{k}: {rutas[k]}" for k in obligatorias if k not in rutas or not Path(rutas[k]).exists()]
    if faltan:
        detalle = "\n".join(f"    · {f}" for f in faltan)
        raise ArchivoFaltante(f"No se encontraron estos archivos:\n{detalle}")


# ══════════════════════════════════════════════════════════════════
# FUNCIONES DE TEXTO — idénticas a 01_control_calidad_norte.py
# ══════════════════════════════════════════════════════════════════
def limpiar_texto(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a"}:
        return None
    return re.sub(r"\s+", " ", s).strip()


def normalizar_texto(x):
    s = limpiar_texto(x)
    if not s:
        return ""
    s = s.upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = s.replace(".", "")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalizar_nombre_para_buscar(x):
    return normalizar_texto(x)


def normalizar_movimiento(x, default=None):
    s = limpiar_texto(x)
    if not s:
        return default
    n = normalizar_texto(s)
    if "INGRES" in n:
        return "Ingreso"
    if "TRASL" in n:
        return "Traslado"
    if "SALIDA" in n:
        return "Salida"
    return s.strip()


def buscar_columna(df, opciones, obligatorio=False):
    mapa = {normalizar_texto(c): c for c in df.columns}
    for op in opciones:
        key = normalizar_texto(op)
        if key in mapa:
            return mapa[key]
    if obligatorio:
        raise ValueError(f"No encontré ninguna columna entre: {opciones}. Columnas disponibles: {list(df.columns)}")
    return None


def tomar_columna(df, opciones, default=None, obligatorio=False):
    col = buscar_columna(df, opciones, obligatorio=obligatorio)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


# ══════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def leer_excel(path, p, **kwargs):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            return pd.read_excel(path, **kwargs)
        except PermissionError:
            p(f"  ⚠ Bloqueado: {Path(path).name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def leer_fuente(path, hoja, p, headers_fallback=(0, 7, 8, 9)):
    path = Path(path)
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
            if hoja not in wb.sheetnames:
                wb.close()
                raise ValueError(f"No existe la hoja '{hoja}' en {path.name}. Hojas: {wb.sheetnames}")
            ws = wb[hoja]
            if ws.tables:
                tabla = list(ws.tables.values())[0]
                filas = [[c.value for c in fila] for fila in ws[tabla.ref]]
                headers = filas[0]
                df = pd.DataFrame(filas[1:], columns=headers)
                wb.close()
                df.columns = df.columns.astype(str).str.strip()
                return df.dropna(how="all").reset_index(drop=True)
            wb.close()

            mejor = None
            mejor_score = -1
            claves = [
                "FECHA", "CLIENTE", "RUT", "GENERADOR", "TRANSPORTISTA",
                "DESTINO", "TIPO", "MOVIMIENTO", "PESO", "TICKET", "PATENTE",
            ]
            for header in headers_fallback:
                try:
                    cand = pd.read_excel(path, sheet_name=hoja, header=header)
                    cand.columns = cand.columns.astype(str).str.strip()
                    cand = cand.dropna(how="all").reset_index(drop=True)
                    score = 0
                    cols_norm = [normalizar_texto(c) for c in cand.columns]
                    for clave in claves:
                        if any(clave in c for c in cols_norm):
                            score += 1
                    if score > mejor_score:
                        mejor = cand
                        mejor_score = score
                except Exception:
                    pass
            if mejor is None:
                raise ValueError(f"No se pudo leer {path.name} / hoja {hoja}")
            return mejor
        except PermissionError:
            p(f"  ⚠ Bloqueado: {path.name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def cargar_lookups(rutas, p):
    p("\n── Cargando maestros compartidos desde OneDrive ──")

    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    df_t = df_t[["Transportista", "RUT"]].dropna(subset=["Transportista"])
    df_t["Transportista"] = df_t["Transportista"].astype(str).str.strip()
    df_t["RUT"] = df_t["RUT"].astype(str).str.strip()
    transportistas = dict(zip(df_t["Transportista"], df_t["RUT"]))
    p(f"  Transportistas: {len(transportistas)}")

    df_m = leer_excel(rutas["sinader"], p, sheet_name="Clasificación Residuos", header=1)
    df_m.columns = df_m.columns.astype(str).str.strip()
    df_m = df_m[["TIPO", "CÓDIGOS SINADER"]].dropna(subset=["TIPO"])
    df_m["TIPO"] = df_m["TIPO"].astype(str).str.strip()
    df_m["CÓDIGOS SINADER"] = df_m["CÓDIGOS SINADER"].astype(str).str.strip()
    df_m = df_m.drop_duplicates(subset=["TIPO"])
    materiales = dict(zip(df_m["TIPO"], df_m["CÓDIGOS SINADER"]))
    if "ASIMILABLE A DOMICILIARIO" not in materiales:
        materiales["ASIMILABLE A DOMICILIARIO"] = "20 03 01"
    p(f"  SINADER: {len(materiales)} tipos")

    df_d = leer_excel(rutas["destinatarios"], p, sheet_name="NACIONAL", header=1)
    df_d.columns = df_d.columns.astype(str).str.strip()
    cols = [
        "NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", "CÓDIGOS LER",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER",
    ]
    df_d = df_d[cols].dropna(subset=["NOMBRE DE FANTASÍA"])
    for c in cols:
        df_d[c] = df_d[c].astype(str).str.strip()
    p(f"  Destinatarios: {len(df_d)}")

    return transportistas, materiales, df_d


# ══════════════════════════════════════════════════════════════════
# FUENTES
# ══════════════════════════════════════════════════════════════════
def preparar_control(nombre, df, mapa):
    out = pd.DataFrame(index=df.index)
    out["ORIGEN"] = nombre
    for salida, opciones in mapa.items():
        out[salida] = tomar_columna(df, opciones)
    out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce")
    out["AÑO"] = out["Fecha"].dt.year
    out["_desde"] = out["Fecha"] >= pd.Timestamp(FECHA_DESDE_NORTE)
    return out


def cargar_fuentes_control(rutas, p):
    p("\nLeyendo fuentes Zona Norte...")

    recynor_raw    = leer_fuente(rutas["recynor"],    "INGRESOS", p, headers_fallback=(0, 8, 9, 7))
    bo_iquique_raw = leer_fuente(rutas["bo_iquique"], "BBDD",     p, headers_fallback=(0, 7, 8, 9))

    if "irar_arica" in rutas and Path(rutas["irar_arica"]).exists():
        irar_arica_raw = leer_fuente(rutas["irar_arica"], "INGRESOS", p, headers_fallback=(0, 7, 8, 9))
    else:
        p("  ⚠ BBDD IRAR ARICA.xlsx no encontrado — fuente omitida del control")
        irar_arica_raw = pd.DataFrame()

    fuentes = {
        "RECYNOR": preparar_control("RECYNOR", recynor_raw, {
            "Fecha":         ["FECHA", "Fecha"],
            "Cliente":       ["CLIENTE", "Cliente"],
            "RUT":           ["RUT CLIENTE", "RUT"],
            "Generador":     ["GENERADOR", "Generador"],
            "Transportista": ["TRANSPORTISTA", "Transportista"],
            "Destino":       ["DESTINO", "Destino"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["MOVIMIENTO", "Movimiento"],
            "Peso neto":     ["PESO NETO KG", "PESO NETO", "Peso neto"],
        }),
        "BO_IQUIQUE": preparar_control("BO_IQUIQUE", bo_iquique_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["RUT", "RUT CLIENTE"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["Destino", "DESTINO"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto [KG]", "Peso neto", "PESO NETO KG"],
        }),
        "IRAR_ARICA": preparar_control("IRAR_ARICA", irar_arica_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["Rut Cliente", "RUT CLIENTE", "RUT"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["DESTINO", "Destino"],
            "TIPO":          ["Tipo Residuo", "TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto [kg]", "PESO NETO KG", "Peso neto"],
        }),
    }

    for nombre, df in fuentes.items():
        p(f"  {nombre:<18} {len(df):>6} filas | >= {FECHA_DESDE_NORTE}: {df['_desde'].sum()}")

    return fuentes


# ══════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL — esta es la que llama la aplicación
# ══════════════════════════════════════════════════════════════════
def controlar(rutas, ruta_salida=None, mostrar=True):
    """Ejecuta el control de calidad de Norte sobre las fuentes indicadas."""
    p = Registro(mostrar=mostrar)
    p("Iniciando control de calidad NORTE...")

    verificar_rutas(rutas)

    fuentes = cargar_fuentes_control(rutas, p)
    _, materiales, _ = cargar_lookups(rutas, p)

    resumen = []
    partes = []

    for nombre, df in fuentes.items():
        resumen.append({
            "FUENTE": nombre,
            "FILAS_LEIDAS": len(df),
            f"FILAS_DESDE_{FECHA_DESDE_NORTE[:7]}": int(df["_desde"].sum()),
            "FECHA_MIN": df["Fecha"].min(),
            "FECHA_MAX": df["Fecha"].max(),
        })
        partes.append(df[df["_desde"]].copy())

    df_resumen = pd.DataFrame(resumen)
    df_all = pd.concat(partes, ignore_index=True)

    # Movimientos según regla operacional.
    # Recynor: la columna MOVIMIENTO manda. RETIRO equivale a Traslado.
    # BO Iquique: siempre Traslado.
    df_all["Movimiento_norm"] = df_all["Movimiento"].apply(
        lambda x: "Traslado" if "RETIRO" in normalizar_texto(x) else normalizar_movimiento(x)
    )
    df_all.loc[df_all["ORIGEN"].eq("BO_IQUIQUE"), "Movimiento_norm"] = "Traslado"
    # IRAR Arica: planta + BO. Vacío = Ingreso.
    df_all.loc[
        df_all["ORIGEN"].eq("IRAR_ARICA") & df_all["Movimiento_norm"].isna(),
        "Movimiento_norm"
    ] = "Ingreso"

    # En Recynor (hoja INGRESOS), movimiento vacío se interpreta como Ingreso.
    df_all.loc[
        df_all["ORIGEN"].eq("RECYNOR") & df_all["Movimiento_norm"].isna(),
        "Movimiento_norm"
    ] = "Ingreso"

    # Clientes / RUT
    clientes = df_all[["ORIGEN", "Cliente", "RUT"]].copy()
    clientes["CLIENTE_NORM"] = clientes["Cliente"].apply(normalizar_texto)
    clientes["RUT_NORM"] = clientes["RUT"].apply(lambda x: normalizar_texto(x).replace("-", ""))
    clientes = clientes[(clientes["CLIENTE_NORM"] != "") | (clientes["RUT_NORM"] != "")].drop_duplicates()

    rut_por_nombre = clientes.groupby("CLIENTE_NORM")["RUT_NORM"].nunique().reset_index(name="N_RUT")
    nombres_conf = rut_por_nombre[(rut_por_nombre["CLIENTE_NORM"] != "") & (rut_por_nombre["N_RUT"] > 1)]["CLIENTE_NORM"]
    c1 = clientes[clientes["CLIENTE_NORM"].isin(nombres_conf)].copy()
    if not c1.empty:
        c1.insert(0, "PROBLEMA", "Mismo cliente normalizado con RUT distinto")

    nombre_por_rut = clientes.groupby("RUT_NORM")["CLIENTE_NORM"].nunique().reset_index(name="N_CLIENTE")
    ruts_conf = nombre_por_rut[(nombre_por_rut["RUT_NORM"] != "") & (nombre_por_rut["N_CLIENTE"] > 1)]["RUT_NORM"]
    c2 = clientes[clientes["RUT_NORM"].isin(ruts_conf)].copy()
    if not c2.empty:
        c2.insert(0, "PROBLEMA", "Mismo RUT con nombre distinto")

    # Tipo sin SINADER
    tipos = df_all[["ORIGEN", "TIPO"]].drop_duplicates().copy()
    tipos["TIPO_LIMPIO"] = tipos["TIPO"].astype(str).str.strip()
    c_tipo = tipos[
        ~tipos["TIPO_LIMPIO"].isin(set(materiales.keys()))
        & (~tipos["TIPO_LIMPIO"].str.lower().isin(["nan", "none", ""]))
    ].copy()
    if not c_tipo.empty:
        c_tipo.insert(0, "PROBLEMA", "TIPO sin código SINADER")

    # Transportistas sin RUT
    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    trans_ok = set(df_t["Transportista"].dropna().astype(str).str.strip())

    trans = df_all[["ORIGEN", "Transportista"]].drop_duplicates().copy()
    trans["Transportista"] = trans["Transportista"].apply(limpiar_texto)
    c_trans = trans[(trans["Transportista"].notna()) & (~trans["Transportista"].isin(trans_ok))].copy()
    if not c_trans.empty:
        c_trans.insert(0, "PROBLEMA", "Transportista sin RUT en maestro Transportistas")

    # Vacíos críticos
    criticas = ["Fecha", "Cliente", "RUT", "Generador", "TIPO", "Movimiento_norm", "Peso neto"]
    filas_vacios = []
    for col in criticas:
        serie = df_all[col] if col in df_all.columns else pd.Series([None] * len(df_all))
        vacio = serie.isna() | serie.astype(str).str.strip().str.lower().isin(["", "nan", "none", "nat", "#n/a"])
        if vacio.any():
            por_origen = df_all.loc[vacio].groupby("ORIGEN").size().reset_index(name="CANTIDAD")
            for _, r in por_origen.iterrows():
                filas_vacios.append({
                    "PROBLEMA": "Columna crítica vacía",
                    "COLUMNA": col,
                    "ORIGEN": r["ORIGEN"],
                    "CANTIDAD": int(r["CANTIDAD"]),
                })
    c_vacios = pd.DataFrame(filas_vacios)

    # Destinos vacíos: BO Iquique siempre; Recynor solo en los RETIRO/Traslado.
    destinos_revisar = df_all.copy()
    necesita_destino_archivo = (
        destinos_revisar["ORIGEN"].eq("BO_IQUIQUE")
        | (
            destinos_revisar["ORIGEN"].eq("RECYNOR")
            & destinos_revisar["Movimiento_norm"].eq("Traslado")
        )
    )
    destino_vacio = destinos_revisar["Destino"].isna() | destinos_revisar["Destino"].astype(str).str.strip().str.lower().isin(["", "nan", "none", "nat"])
    c_dest = destinos_revisar.loc[necesita_destino_archivo & destino_vacio, ["ORIGEN", "Fecha", "Cliente", "Movimiento_norm", "Destino"]].copy()
    if not c_dest.empty:
        c_dest.insert(0, "PROBLEMA", "Destino vacío donde corresponde usar columna Destino")

    # C3: Generadores con variantes de nombre entre bases
    generadores = df_all[["ORIGEN", "Generador"]].copy()
    generadores["GEN_NORM"] = generadores["Generador"].apply(normalizar_texto)
    generadores = generadores[generadores["GEN_NORM"] != ""].drop_duplicates()

    nombres_por_gen = generadores.groupby("GEN_NORM")["Generador"].nunique().reset_index(name="N_NOMBRE")
    gens_conf = nombres_por_gen[nombres_por_gen["N_NOMBRE"] > 1]["GEN_NORM"]
    c3 = generadores[generadores["GEN_NORM"].isin(gens_conf)].copy()
    if not c3.empty:
        c3.insert(0, "PROBLEMA", "Generador con variantes de nombre entre bases")

    # Nombres para poblar HOMOLOGACION.xlsx
    cli_homolog = (
        clientes[["ORIGEN", "Cliente", "RUT"]]
        .drop_duplicates().sort_values(["ORIGEN", "Cliente"])
        .rename(columns={"ORIGEN": "PLANTA", "Cliente": "NOMBRE_VARIANTE"})
        .reset_index(drop=True)
    )
    cli_homolog["NOMBRE_CORRECTO"] = ""
    cli_homolog["RUT_CORRECTO"] = ""

    gen_homolog = (
        generadores[["ORIGEN", "Generador"]]
        .drop_duplicates().sort_values(["ORIGEN", "Generador"])
        .rename(columns={"ORIGEN": "PLANTA", "Generador": "NOMBRE_VARIANTE"})
        .reset_index(drop=True)
    )
    gen_homolog["NOMBRE_CORRECTO"] = ""

    # Movimientos raros
    mov_ok = {"Ingreso", "Traslado"}
    c_mov = df_all[
        df_all["Movimiento_norm"].isna() | (~df_all["Movimiento_norm"].isin(mov_ok))
    ][["ORIGEN", "Fecha", "Cliente", "Movimiento", "Movimiento_norm"]].drop_duplicates().copy()
    if not c_mov.empty:
        c_mov.insert(0, "PROBLEMA", "Movimiento vacío o no reconocido")

    total = len(c1) + len(c2) + len(c3) + len(c_tipo) + len(c_trans) + len(c_vacios) + len(c_mov) + len(c_dest)

    res = {
        "df_resumen": df_resumen,
        "c1": c1, "c2": c2, "c3": c3,
        "c_tipo": c_tipo, "c_trans": c_trans, "c_vacios": c_vacios,
        "c_mov": c_mov, "c_dest": c_dest,
        "cli_homolog": cli_homolog, "gen_homolog": gen_homolog,
        "total": total,
    }

    if ruta_salida:
        escribir_excel(res, ruta_salida, p)

    res["log"] = p.texto()
    return res


def escribir_excel(res, ruta_salida, p):
    p(f"\nGuardando {Path(ruta_salida).name}...")

    def hoja(writer, df, nombre, msg_ok):
        out = df if not df.empty else pd.DataFrame({"resultado": [msg_ok]})
        out.to_excel(writer, sheet_name=nombre, index=False)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        res["df_resumen"].to_excel(writer, sheet_name="RESUMEN", index=False)
        hoja(writer, res["c1"], "C1_CLIENTE_RUT", "Sin conflictos")
        hoja(writer, res["c2"], "C2_RUT_CLIENTE", "Sin conflictos")
        hoja(writer, res["c3"], "C3_GENERADORES", "Sin variantes")
        res["cli_homolog"].to_excel(writer, sheet_name="HOMOLOG_clientes", index=False)
        res["gen_homolog"].to_excel(writer, sheet_name="HOMOLOG_generadores", index=False)
        hoja(writer, res["c_tipo"], "TIPO_SIN_SINADER", "Todos los TIPO tienen SINADER")
        hoja(writer, res["c_trans"], "TRANSPORT_SIN_RUT", "Todos los transportistas tienen RUT")
        hoja(writer, res["c_vacios"], "VACIOS_CRITICOS", "Sin vacíos críticos")
        hoja(writer, res["c_mov"], "MOVIMIENTOS_REVISAR", "Movimientos OK")
        hoja(writer, res["c_dest"], "DESTINOS_REVISAR", "Sin destinos vacíos críticos")

    # Mensaje idéntico al original (dice "SUR" — ver Hallazgo 6 en el docstring del módulo).
    p("✓ Control de calidad SUR generado")
    p(f"  Archivo: {ruta_salida}")


# ══════════════════════════════════════════════════════════════════
# USO DESDE LA TERMINAL
# ══════════════════════════════════════════════════════════════════
def _main():
    argumentos = [a for a in sys.argv[1:] if not a.startswith("--")]

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    try:
        from config import carpetas, ConfiguracionFaltante
    except ImportError:
        print("\n✗ No se encontró config.py. Debe estar en la carpeta src/.\n")
        sys.exit(2)

    try:
        if argumentos:
            rutas = rutas_desde_carpeta(argumentos[0])
            salida = Path(argumentos[0]) / "CONTROL_CALIDAD_NORTE.xlsx"
        else:
            rutas = rutas_desde_config()
            salida = carpetas()["norte"] / "CONTROL_CALIDAD_NORTE.xlsx"
    except ConfiguracionFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)
    except ArchivoFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    sello = datetime.now().strftime("%Y-%m-%d_%H%M")
    salida = salida.with_name(f"{salida.stem}_{sello}{salida.suffix}")

    try:
        res = controlar(rutas, ruta_salida=salida)
    except ArchivoFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)
    except ValueError as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    sys.exit(0 if res["total"] == 0 else 1)


if __name__ == "__main__":
    _main()
