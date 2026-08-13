"""
consolidar.py — Consolidador de Trazabilidad Zona Sur (versión refactorizada)
Ambipar Group — Proyecto TRZ-APP-001, Etapa 1

Mismo patrón que src/rm/consolidar.py y src/sur/control_calidad.py: la lógica
de las 9 fuentes, el cruce con SINADER y el cruce (de tres pasadas) con
Destinatarios son exactamente los mismos que en 02_consolidar_sur.py, línea
por línea. Lo que cambia es que las rutas se reciben como parámetro y el log
de auditoría ya no se abre automáticamente al importar el archivo.

REGLAS DEFINITIVAS DESTINO ZONA SUR (documentadas en el script original, sin
cambios — ver también CONTROL_CALIDAD_SUR y el documento TRZ-APP-001):
    - NO se infieren destinos por histórico. Destino inferido = No siempre.
    - Coronel / San Joaquín: Ingreso → planta; Traslado → columna DESTINO.
    - BO Trapén: SIEMPRE Traslado.
    - Ecofibras Planta Trapén: NO maneja Movimiento → TODO Ingreso.
    - Temuco: todos Traslado.
    - Proveedores externos (MOWI): Movimiento = "Traslado externo".
    - BO Chiloé: solo traslados; se EXCLUYE el cliente MOWI.
    - I.R.A.R Los Ángeles: columna MOVIMIENTO manda (Ingreso → planta,
      Traslado → columna DESTINO). No está en PLANTAS_PROPIAS_ACTIVAS.
    - BO Chillán: siempre Traslado, Región de Ñuble.

Se puede usar de dos formas:

    Desde otro programa:
        from consolidar import consolidar, rutas_desde_config
        rutas = rutas_desde_config()
        resultado = consolidar(rutas, modo_prueba=True)

    Desde la terminal, igual que antes:
        python consolidar.py --prueba
        python consolidar.py --reset
"""

import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook


# ══════════════════════════════════════════════════════════════════
# CONSTANTES DE LA ZONA SUR — reglas de negocio, no rutas
# ══════════════════════════════════════════════════════════════════
FECHA_DESDE_SUR = "2026-04-01"
HOJA_DESTINO = "TRAZABILIDAD"

REGION_CORONEL       = "Región del Biobío"
REGION_BO_TRAPEN     = "Región de Los Lagos"
REGION_PLANTA_TRAPEN = "Región de Los Lagos"
REGION_SAN_JOAQUIN   = "Región de Los Lagos"
REGION_IRAR          = "Región del Biobío"
REGION_TEMUCO        = "Región de la Araucanía"
REGION_BO_CHILLAN    = "Región de Ñuble"
REGION_BO_CHILOE     = "Región de Los Lagos"
REGION_PROVEEDORES   = "Región de Los Lagos"

GESTOR_AMBIPAR = "AMBIPAR ENVIRONMENT CHILE"
RUT_AMBIPAR    = "96824110-9"

HOJA_DESTINATARIOS = "NACIONAL"

# Plantas propias que ya están en el pipeline. Ver docstring del script
# original para la explicación completa de por qué IRAR queda fuera.
PLANTAS_PROPIAS_ACTIVAS = {
    "ECOFIBRAS CORONEL",
    "ECOFIBRAS PLANTA TRAPÉN",
    "ECOFIBRAS SAN JOAQUÍN",
    "ECOFIBRAS PUERTO MONTT",
    "ECOFIBRAS TRAPÉN",
}

COLUMNAS_FINALES = [
    "Fecha", "Mes", "Cliente", "RUT", "Gestor", "Contrato", "Generador",
    "Transportista", "Rut transportista", "Patente de Camión", "Ticket de pesaje",
    "Peso neto (kg)", "Destino", "Comuna Destino", "RUT DESTINATARIO",
    "TIPO", "CÓDIGOS SINADER", "Movimiento", "Movimiento interempresa",
    "CÓDIGO ESTABLECIMIENTO SINADER", "CÓDIGO DE TRATAMIENTO SINADER",
    "Región", "Destino inferido",
]

CLAVE_DEDUP = ["Fecha", "Cliente", "Ticket de pesaje", "TIPO", "Destino"]

MAX_REINTENTOS = 5
ESPERA_REINTENTO = 30

MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


class ArchivoFaltante(Exception):
    """Se lanza cuando falta algún archivo fuente obligatorio."""


# ══════════════════════════════════════════════════════════════════
# REGISTRO DE MENSAJES — reemplaza al logging.basicConfig de nivel de módulo
# ══════════════════════════════════════════════════════════════════
class Registro:
    """Ver docstring de la clase equivalente en src/rm/consolidar.py."""

    def __init__(self, mostrar=True, ruta_log=None):
        self.lineas = []
        self.mostrar = mostrar
        self.ruta_log = Path(ruta_log) if ruta_log else None

    def __call__(self, msg=""):
        msg = str(msg)
        self.lineas.append(msg)
        if self.mostrar:
            print(msg)
        if self.ruta_log:
            marca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(self.ruta_log, "a", encoding="utf-8") as f:
                f.write(f"{marca}  {msg}\n")

    def texto(self):
        return "\n".join(self.lineas)


# ══════════════════════════════════════════════════════════════════
# CÓMO ENCONTRAR LOS ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def rutas_desde_config(base=None):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas

    c = carpetas(base)
    zona = c["zona_sur"]
    info = c["info_base"]

    rutas = {
        "coronel":        zona / "BBDD ECOFIBRAS CORONEL.xlsx",
        "bo_trapen":       zona / "BBDD BO TRAPEN.xlsx",
        "planta_trapen":   zona / "BBDD ECOFIBRAS PLANTA TRAPEN .xlsx",
        "san_joaquin":     zona / "BBDD ECOFIBRAS SAN JOAQUIN.xlsx",
        "irar":            zona / "BBDD IRAR LOS ÁNGELES.xlsx",
        "bo_chillan":      zona / "BBDD BO CHILLÁN.xlsx",
        "temuco":          zona / "BBDD TEMUCO.xlsx",
        "proveedores":     zona / "BBDD PROVEEDORES MOWI.xlsx",
        "bo_chiloe":       zona / "BBDD BO CHILOÉ.xlsx",
        "homologacion":    info / "HOMOLOGACION.xlsx",
        "destinatarios":   info / "BBDD_DESTINATARIO.xlsx",
        "sinader":         info / "Clasificación_Residuos SINADER.xlsx",
        "transportistas":  info / "Transportistas.xlsx",
        "destino_real":    c["bbdd"] / "BBDD_TRAZABILIDAD_SUR.xlsx",
    }
    alt_irar = zona / "BBDD I.R.A.R LOS ÁNGELES.xlsx"
    if not rutas["irar"].exists() and alt_irar.exists():
        rutas["irar"] = alt_irar
    alt_dest = info / "BBDD DESTINATARIO.xlsx"
    if not rutas["destinatarios"].exists() and alt_dest.exists():
        rutas["destinatarios"] = alt_dest
    alt_prov = zona / "BBDD PROVEEDORES.xlsx"
    if not rutas["proveedores"].exists() and alt_prov.exists():
        rutas["proveedores"] = alt_prov
    alt_chiloe = zona / "BBDD BO CHILOE.xlsx"
    if not rutas["bo_chiloe"].exists() and alt_chiloe.exists():
        rutas["bo_chiloe"] = alt_chiloe

    return rutas


def verificar_rutas(rutas):
    """proveedores y bo_chiloe son opcionales — el script original los omite
    si no existen, en vez de fallar (ver procesar_proveedores/procesar_bo_chiloe)."""
    obligatorias = [
        "coronel", "bo_trapen", "planta_trapen", "san_joaquin", "irar",
        "bo_chillan", "temuco", "homologacion", "destinatarios",
        "sinader", "transportistas",
    ]
    faltan = [f"{k}: {rutas[k]}" for k in obligatorias if k not in rutas or not Path(rutas[k]).exists()]
    if faltan:
        detalle = "\n".join(f"    · {f}" for f in faltan)
        raise ArchivoFaltante(f"No se encontraron estos archivos:\n{detalle}")


# ══════════════════════════════════════════════════════════════════
# FUNCIONES DE TEXTO — idénticas a 02_consolidar_sur.py
# ══════════════════════════════════════════════════════════════════
def nombre_mes(fecha):
    try:
        return MESES_ES[pd.Timestamp(fecha).month]
    except Exception:
        return ""


def limpiar_texto(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a"}:
        return None
    return re.sub(r"\s+", " ", s).strip()


def normalizar_texto(x):
    s = limpiar_texto(x)
    if not s:
        return ""
    s = s.upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = s.replace(".", "")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalizar_nombre_para_buscar(x):
    return normalizar_texto(x)


def normalizar_region_para_cruce(x):
    """Normaliza regiones para poder cruzar (ej. 'Región de Los Lagos' -> 'LOS LAGOS')."""
    n = normalizar_texto(x)
    n = n.replace("REGION DE ", "")
    n = n.replace("REGION DEL ", "")
    n = n.replace("REGION ", "")
    n = n.replace("METROPOLITANA DE SANTIAGO", "METROPOLITANA")
    return n.strip()


def normalizar_ler(x):
    """Códigos LER/SINADER comparables: '17 01 07', '17-01-07' y '170107' son iguales."""
    return re.sub(r"[^A-Z0-9]", "", normalizar_texto(x))


def es_cliente_resimple(x):
    """Detecta RESIMPLE aunque venga como RE SIMPLE, RE-SIMPLE, RESIMPLE S.A., etc."""
    n = normalizar_texto(x)
    n_sin_espacios = re.sub(r"[^A-Z0-9]", "", n)
    return "RESIMPLE" in n_sin_espacios


def es_cliente_mowi(x):
    """True si el cliente corresponde a MOWI (contrato excluido en Chiloé)."""
    n = normalizar_texto(x)
    return bool(n) and "MOWI" in n


def limpiar_numero(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a", "-"}:
        return None
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
        s = s.replace(".", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", ".", "-", "-."}:
        return None
    try:
        return float(s)
    except Exception:
        return None


def normalizar_movimiento(x, default=None):
    s = limpiar_texto(x)
    if not s:
        return default
    n = normalizar_texto(s)
    if "INGRES" in n:
        return "Ingreso"
    if "TRASL" in n:
        return "Traslado"
    if "SALIDA" in n:
        return "Salida"
    return s.strip()


def normalizar_destino(x, planta_si_ecofibras=None):
    s = limpiar_texto(x)
    if not s:
        return None
    n = normalizar_texto(s)
    if n in {"ECOFIBRAS", "ECOFIBRAS SA", "ECOFIBRAS S A"} and planta_si_ecofibras:
        return planta_si_ecofibras
    if "ECOFIBRAS CORONEL" in n:
        return "ECOFIBRAS CORONEL"
    if "BO TRAPEN" in n:
        return "BO TRAPÉN"
    if "ECOFIBRAS PLANTA TRAPEN" in n:
        return "ECOFIBRAS PLANTA TRAPÉN"
    if "ECOFIBRAS TRAPEN" in n or "ECOFIBRAS TRAP" in n:
        return "ECOFIBRAS TRAPÉN"
    if "ECOFIBRAS SAN JOAQUIN" in n:
        return "ECOFIBRAS SAN JOAQUÍN"
    if "ECOFIBRAS PUERTO MONTT" in n:
        return "ECOFIBRAS PUERTO MONTT"
    if "IRAR ARICA" in n or ("IRAR" in n and "ARICA" in n):
        return "IRAR ARICA"
    if "I R A R" in n or "IRAR" in n:
        return "IRAR LOS ÁNGELES"
    if "LOS ANGELES" in n and "IRAR" in n:
        return "IRAR LOS ÁNGELES"
    if "RECYNOR" in n:
        return "RECYNOR"
    if "GIRI" in n:
        return "GIRI"
    if "REXIN" in n:
        return "REXIN"
    if n == "BDC":
        return "BDC"
    return s.strip().upper()


def comuna_por_destino(destino):
    d = normalizar_texto(destino)
    if not d:
        return None
    if "ECOFIBRAS CORONEL" in d:
        return "CORONEL"
    if "BO TRAPEN" in d:
        return "PUERTO MONTT"
    if "ECOFIBRAS PLANTA TRAPEN" in d:
        return "PUERTO MONTT"
    if "ECOFIBRAS TRAPEN" in d or "TRAPEN" in d:
        return "PUERTO MONTT"
    if "ECOFIBRAS SAN JOAQUIN" in d or "SAN JOAQUIN" in d:
        return "PUERTO MONTT"
    if "ECOFIBRAS PUERTO MONTT" in d:
        return "PUERTO MONTT"
    if "IRAR" in d or "I R A R" in d or "LOS ANGELES" in d:
        return "LOS ÁNGELES"
    if d == "BDC":
        return "LAUTARO"
    if "GIRI" in d:
        return "QUILICURA"
    return None


def normalizar_gestor_ambipar(x):
    """Unifica cualquier variante de AMBIPAR al nombre oficial del gestor."""
    s = limpiar_texto(x)
    if s is None:
        return None
    return GESTOR_AMBIPAR if "AMBIPAR" in normalizar_texto(s) else s


def buscar_columna(df, opciones, obligatorio=False):
    mapa = {normalizar_texto(c): c for c in df.columns}
    for op in opciones:
        key = normalizar_texto(op)
        if key in mapa:
            return mapa[key]
    if obligatorio:
        raise ValueError(f"No encontré ninguna columna entre: {opciones}. Columnas disponibles: {list(df.columns)}")
    return None


def tomar_columna(df, opciones, default=None, obligatorio=False):
    col = buscar_columna(df, opciones, obligatorio=obligatorio)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


# ══════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def leer_excel(path, p, **kwargs):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            return pd.read_excel(path, **kwargs)
        except PermissionError:
            p(f"  ⚠ Bloqueado: {Path(path).name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def leer_fuente(path, hoja, p, headers_fallback=(0, 7, 8, 9)):
    path = Path(path)
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
            if hoja not in wb.sheetnames:
                wb.close()
                raise ValueError(f"No existe la hoja '{hoja}' en {path.name}. Hojas: {wb.sheetnames}")
            ws = wb[hoja]
            if ws.tables:
                tabla = list(ws.tables.values())[0]
                filas = [[c.value for c in fila] for fila in ws[tabla.ref]]
                headers = filas[0]
                df = pd.DataFrame(filas[1:], columns=headers)
                wb.close()
                df.columns = df.columns.astype(str).str.strip()
                return df.dropna(how="all").reset_index(drop=True)
            wb.close()

            mejor = None
            mejor_score = -1
            claves = [
                "FECHA", "CLIENTE", "RUT", "GENERADOR", "TRANSPORTISTA",
                "DESTINO", "TIPO", "MOVIMIENTO", "PESO", "TICKET", "PATENTE",
            ]
            for header in headers_fallback:
                try:
                    cand = pd.read_excel(path, sheet_name=hoja, header=header)
                    cand.columns = cand.columns.astype(str).str.strip()
                    cand = cand.dropna(how="all").reset_index(drop=True)
                    score = 0
                    cols_norm = [normalizar_texto(c) for c in cand.columns]
                    for clave in claves:
                        if any(clave in c for c in cols_norm):
                            score += 1
                    if score > mejor_score:
                        mejor = cand
                        mejor_score = score
                except Exception:
                    pass
            if mejor is None:
                raise ValueError(f"No se pudo leer {path.name} / hoja {hoja}")
            return mejor
        except PermissionError:
            p(f"  ⚠ Bloqueado: {path.name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def leer_tabla_nombrada(path, nombre_tabla, p):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            wb = load_workbook(path, data_only=True)
            for sheet in wb.worksheets:
                for tabla in sheet.tables.values():
                    if tabla.name == nombre_tabla:
                        filas = [[c.value for c in fila] for fila in sheet[tabla.ref]]
                        headers = filas[0]
                        df = pd.DataFrame(filas[1:], columns=headers)
                        wb.close()
                        return df.dropna(how="all").reset_index(drop=True)
            wb.close()
            raise ValueError(f"Tabla '{nombre_tabla}' no encontrada en {Path(path).name}")
        except PermissionError:
            p(f"  ⚠ Bloqueado: {Path(path).name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {Path(path).name}")


def cargar_homologacion(ruta_homologacion, p):
    homolog_clientes = {}
    homolog_generadores = {}

    ruta = Path(ruta_homologacion)
    p("\n── Cargando HOMOLOGACION desde OneDrive ──")
    p(f"  {ruta}")

    if not ruta.exists():
        p("  ⚠ No se encontró HOMOLOGACION.xlsx — se continúa sin correcciones")
        return homolog_clientes, homolog_generadores

    try:
        df_cli = leer_excel(ruta, p, sheet_name="clientes")
        df_cli = df_cli.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_cli.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            rut_correcto = limpiar_texto(row.get("RUT_CORRECTO"))
            if variante and nombre_correcto:
                homolog_clientes[variante] = {"nombre": nombre_correcto, "rut": rut_correcto}
        p(f"  Clientes homologados: {len(homolog_clientes)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja clientes: {e}")

    try:
        df_gen = leer_excel(ruta, p, sheet_name="generadores")
        df_gen = df_gen.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_gen.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            if variante and nombre_correcto:
                homolog_generadores[variante] = nombre_correcto
        p(f"  Generadores homologados: {len(homolog_generadores)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja generadores: {e}")

    return homolog_clientes, homolog_generadores


def aplicar_homologacion_cliente(df, col_nombre, col_rut, homolog_clientes):
    if not homolog_clientes or col_nombre not in df.columns:
        return df
    if col_rut not in df.columns:
        df[col_rut] = None

    def corregir(row):
        variante = normalizar_nombre_para_buscar(row.get(col_nombre))
        if variante in homolog_clientes:
            corr = homolog_clientes[variante]
            nuevo_nombre = corr["nombre"]
            nuevo_rut = corr["rut"] if corr["rut"] else row.get(col_rut)
            return pd.Series([nuevo_nombre, nuevo_rut])
        return pd.Series([row.get(col_nombre), row.get(col_rut)])

    df[[col_nombre, col_rut]] = df.apply(corregir, axis=1)
    return df


def aplicar_homologacion_generador(df, col_generador, homolog_generadores):
    if not homolog_generadores or col_generador not in df.columns:
        return df

    def corregir(x):
        key = normalizar_nombre_para_buscar(x)
        return homolog_generadores.get(key, x)

    df[col_generador] = df[col_generador].apply(corregir)
    return df


def cargar_lookups(rutas, p):
    p("\n── Cargando maestros compartidos desde OneDrive ──")

    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    df_t = df_t[["Transportista", "RUT"]].dropna(subset=["Transportista"])
    df_t["Transportista"] = df_t["Transportista"].astype(str).str.strip()
    df_t["RUT"] = df_t["RUT"].astype(str).str.strip()
    transportistas = dict(zip(df_t["Transportista"], df_t["RUT"]))
    p(f"  Transportistas: {len(transportistas)}")

    df_m = leer_excel(rutas["sinader"], p, sheet_name="Clasificación Residuos", header=1)
    df_m.columns = df_m.columns.astype(str).str.strip()
    df_m = df_m[["TIPO", "CÓDIGOS SINADER"]].dropna(subset=["TIPO"])
    df_m["TIPO"] = df_m["TIPO"].astype(str).str.strip()
    df_m["CÓDIGOS SINADER"] = df_m["CÓDIGOS SINADER"].astype(str).str.strip()
    df_m = df_m.drop_duplicates(subset=["TIPO"])
    materiales = dict(zip(df_m["TIPO"], df_m["CÓDIGOS SINADER"]))
    if "ASIMILABLE A DOMICILIARIO" not in materiales:
        materiales["ASIMILABLE A DOMICILIARIO"] = "20 03 01"
    p(f"  SINADER: {len(materiales)} tipos")

    df_d = leer_excel(rutas["destinatarios"], p, sheet_name=HOJA_DESTINATARIOS, header=1)
    df_d.columns = df_d.columns.astype(str).str.strip()

    col_region_dest = buscar_columna(
        df_d,
        ["REGIÓN DE ESTABLECIMIETNO", "REGIÓN DE ESTABLECIMIENTO", "REGION DE ESTABLECIMIENTO", "REGIÓN", "REGION"],
        obligatorio=True,
    )

    REGIONES_EXCLUIR = {"METROPOLITANA", "REGION METROPOLITANA", "RM"}

    n_dest_antes = len(df_d)
    df_d["_REGION_DEST_NORM"] = df_d[col_region_dest].apply(normalizar_region_para_cruce)
    df_d = df_d[~df_d["_REGION_DEST_NORM"].isin(REGIONES_EXCLUIR)].copy()
    p(f"  Destinatarios filtrados (excl. RM): {len(df_d)} de {n_dest_antes}")

    cols = [
        "NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", col_region_dest, "CÓDIGOS LER",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER",
    ]
    df_d = df_d[cols].dropna(subset=["NOMBRE DE FANTASÍA"])
    for c in cols:
        df_d[c] = df_d[c].astype(str).str.strip()
    df_d = df_d.rename(columns={col_region_dest: "REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"})

    p(f"  Destinatarios SUR disponibles para cruce LER+Región: {len(df_d)}")

    return transportistas, materiales, df_d


def guardar_excel(df, ruta, p):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            if ruta.exists():
                wb = load_workbook(ruta)
                if HOJA_DESTINO in wb.sheetnames:
                    del wb[HOJA_DESTINO]
                ws = wb.create_sheet(HOJA_DESTINO)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = HOJA_DESTINO
            ws.append(list(df.columns))
            for fila in df.itertuples(index=False):
                ws.append([None if str(v) in ("nan", "NaT", "None") else v for v in fila])
            wb.save(ruta)
            p(f"  ✓ Guardado en: {ruta}")
            return
        except PermissionError:
            p(f"  ⚠ Archivo bloqueado — intento {intento}/{MAX_REINTENTOS}: {ruta.name}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo guardar {ruta}. Ciérralo en Excel.")


# ══════════════════════════════════════════════════════════════════
# FUENTES
# ══════════════════════════════════════════════════════════════════
def base_final(n):
    return pd.DataFrame(index=range(n), columns=[
        "Fecha", "Mes", "Cliente", "RUT", "Gestor", "Contrato", "Generador",
        "Transportista", "Rut transportista", "Patente de Camión", "Ticket de pesaje",
        "Peso neto", "Unidad", "Destino", "Comuna Destino", "TIPO", "Movimiento",
        "Movimiento interempresa", "Región", "Destino inferido",
    ])


def destino_coronel(row):
    destino_excel = normalizar_destino(row.get("DESTINO_COL"), planta_si_ecofibras="ECOFIBRAS CORONEL")
    if destino_excel:
        return destino_excel
    if row.get("Movimiento_final") == "Ingreso":
        return "ECOFIBRAS CORONEL"
    return None


def destino_planta_trapen(row):
    destino_excel = normalizar_destino(row.get("DESTINO_COL"), planta_si_ecofibras="ECOFIBRAS PLANTA TRAPÉN")
    if destino_excel:
        return destino_excel
    if row.get("Movimiento_final") == "Ingreso":
        return "ECOFIBRAS PLANTA TRAPÉN"
    return None


def destino_san_joaquin(row):
    destino_excel = normalizar_destino(row.get("DESTINO_COL"), planta_si_ecofibras="ECOFIBRAS SAN JOAQUÍN")
    if destino_excel:
        return destino_excel
    if row.get("Movimiento_final") == "Ingreso":
        return "ECOFIBRAS SAN JOAQUÍN"
    return None


def destino_irar(row):
    destino_excel = normalizar_destino(row.get("DESTINO_COL"))
    if row.get("Movimiento_final") == "Ingreso":
        return "IRAR LOS ÁNGELES"
    return destino_excel


def filtrar_cliente_resimple(df, col_cliente, nombre_fuente, p):
    if col_cliente not in df.columns:
        p(f"  ⚠ No se encontró columna cliente para filtrar RESIMPLE en {nombre_fuente}")
        return df
    mask_resimple = df[col_cliente].apply(es_cliente_resimple)
    n = int(mask_resimple.sum())
    if n > 0:
        p(f"  • Filtro Cliente contiene RESIMPLE: excluidas {n} filas en {nombre_fuente}")
    return df.loc[~mask_resimple].copy()


def excluir_traslados_a_planta_propia(out, nombre_fuente, p):
    if "Destino" not in out.columns or out.empty:
        return out
    plantas_norm = {normalizar_texto(pl) for pl in PLANTAS_PROPIAS_ACTIVAS}
    dest_norm = out["Destino"].apply(normalizar_texto)
    es_traslado = out["Movimiento"].astype(str).str.strip().str.lower().ne("ingreso")
    mask = dest_norm.isin(plantas_norm) & es_traslado
    n = int(mask.sum())
    if n > 0:
        detalle = out.loc[mask].groupby("Destino").size().to_dict()
        p(f"  • Traslados a planta propia descartados en {nombre_fuente}: {n}")
        for destino, cnt in detalle.items():
            p(f"     → {destino}: {cnt} viajes (los registra la planta como Ingreso)")
        p(f"     Si la planta no los registró, esos {n} viajes NO quedan trazados.")
    return out.loc[~mask].copy()


def _procesar_ecofibras(
    rutas, clave_ruta, nombre_log, planta_nombre, region,
    fn_destino, transportistas, homolog_c, homolog_g, p,
    forzar_ingreso=False,
):
    p(f"\n── {nombre_log} ──")

    df = leer_fuente(rutas[clave_ruta], "INGRESOS", p, headers_fallback=(8, 9, 7, 0))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["FECHA", "Fecha"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    mov_col    = tomar_columna(df, ["MOVIMIENTO", "Movimiento"])
    destino_col= tomar_columna(df, ["DESTINO", "Destino", "DESTINO FINAL", "Destino Final"])

    tmp = pd.DataFrame(index=df.index)
    tmp["Movimiento_final"] = mov_col.apply(lambda x: normalizar_movimiento(x, default=None))
    tmp["DESTINO_COL"]      = destino_col

    col_dest_encontrada = buscar_columna(df, ["DESTINO", "Destino", "DESTINO FINAL", "Destino Final"])
    p(f"  • Columna DESTINO usada: {col_dest_encontrada!r}")
    dist_mov = tmp["Movimiento_final"].value_counts(dropna=False).to_dict()
    p(f"  • Distribución Movimiento: {dist_mov}")

    if forzar_ingreso:
        n_traslado = int(tmp["Movimiento_final"].eq("Traslado").sum())
        if n_traslado:
            p(f"  ⚠ {n_traslado} filas traen Movimiento = Traslado; la regla "
              f"'{nombre_log} = solo Ingresos' las registrará igual como Ingreso. REVISAR.")
        dest_raros = (
            tmp["DESTINO_COL"].apply(lambda x: normalizar_destino(x, planta_si_ecofibras=planta_nombre))
            .dropna().loc[lambda s: s != planta_nombre].unique().tolist()
        )
        if dest_raros:
            p(f"  ⚠ Columna DESTINO trae valores distintos de la planta: {dest_raros}. "
              f"Se ignoran; Destino = {planta_nombre}. REVISAR.")
        tmp["Movimiento_final"] = "Ingreso"
        p(f"  • Regla aplicada: TODO Ingreso → Destino = {planta_nombre}")

    out["Fecha"]          = df["_Fecha"].values
    out["Mes"]            = out["Fecha"].apply(nombre_mes)
    out["Cliente"]        = tomar_columna(df, ["CLIENTE", "Cliente"]).apply(limpiar_texto).values
    out["RUT"]            = tomar_columna(df, ["RUT CLIENTE", "RUT"]).apply(limpiar_texto).values
    out["Gestor"]         = tomar_columna(df, ["TRANSPORTISTA", "Transportista", "GESTOR"]).apply(limpiar_texto).values
    out["Contrato"]       = None
    out["Generador"]      = tomar_columna(df, ["GENERADOR", "Generador"]).apply(limpiar_texto).values
    out["Transportista"]  = tomar_columna(df, ["TRANSPORTISTA", "Transportista", "GESTOR"]).apply(limpiar_texto).values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Patente de Camión"] = tomar_columna(df, ["PATENTE CAMIÓN", "PATENTE", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["TICKET DE PESAJE", "Ticket de pesaje"]).apply(limpiar_texto).values
    out["Peso neto"]         = tomar_columna(df, ["PESO NETO KG", "Peso neto", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Movimiento"]        = tmp["Movimiento_final"].values
    if forzar_ingreso:
        out["Destino"]       = planta_nombre
    else:
        out["Destino"]       = tmp.apply(fn_destino, axis=1).values
    out["Comuna Destino"]    = out["Destino"].apply(comuna_por_destino)
    out["TIPO"]              = tomar_columna(df, ["TIPO RESIDUO", "TIPO"]).apply(limpiar_texto).values
    out["Región"]            = region
    out["Destino inferido"]  = "No"

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ {nombre_log} final: {len(out)} filas")
    return out


def procesar_ecofibras_coronel(rutas, transportistas, homolog_c, homolog_g, p):
    return _procesar_ecofibras(
        rutas, "coronel", "Ecofibras Coronel", "ECOFIBRAS CORONEL", REGION_CORONEL,
        destino_coronel, transportistas, homolog_c, homolog_g, p,
    )


def procesar_bo_trapen(rutas, transportistas, homolog_c, homolog_g, p):
    p("\n── BO Trapén — SIEMPRE Traslado (hoja MOVIMIENTOS BO) ──")

    df = leer_fuente(rutas["bo_trapen"], "MOVIMIENTOS BO", p, headers_fallback=(8, 9, 7, 0))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["FECHA", "Fecha"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    col_dest_encontrada = buscar_columna(df, ["DESTINO", "Destino"])
    p(f"  • Columna DESTINO usada: {col_dest_encontrada!r}")

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["CLIENTE", "Cliente"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["RUT CLIENTE", "RUT"]).apply(limpiar_texto).values
    out["Gestor"]            = tomar_columna(df, ["TRANSPORTISTA", "Transportista", "GESTOR"]).apply(limpiar_texto).values
    out["Contrato"]          = None
    out["Generador"]         = tomar_columna(df, ["GENERADOR", "Generador"]).apply(limpiar_texto).values
    out["Transportista"]     = tomar_columna(df, ["TRANSPORTISTA", "Transportista", "GESTOR"]).apply(limpiar_texto).values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Patente de Camión"] = tomar_columna(df, ["PATENTE CAMIÓN", "PATENTE", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["TICKET DE PESAJE", "Ticket de pesaje"]).apply(limpiar_texto).values
    out["Peso neto"]         = tomar_columna(df, ["PESO NETO KG", "Peso neto", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Movimiento"]        = "Traslado"
    out["Destino"]           = tomar_columna(df, ["DESTINO", "Destino"]).apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]    = out["Destino"].apply(comuna_por_destino)
    out["TIPO"]              = tomar_columna(df, ["TIPO RESIDUO", "TIPO"]).apply(limpiar_texto).values
    out["Región"]            = REGION_BO_TRAPEN
    out["Destino inferido"]  = "No"

    out = excluir_traslados_a_planta_propia(out, "BO Trapén", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ BO Trapén final: {len(out)} filas")
    return out


def procesar_ecofibras_planta_trapen(rutas, transportistas, homolog_c, homolog_g, p):
    return _procesar_ecofibras(
        rutas, "planta_trapen", "Ecofibras Planta Trapén — SIEMPRE Ingreso", "ECOFIBRAS PLANTA TRAPÉN", REGION_PLANTA_TRAPEN,
        destino_planta_trapen, transportistas, homolog_c, homolog_g, p,
        forzar_ingreso=True,
    )


def procesar_ecofibras_san_joaquin(rutas, transportistas, homolog_c, homolog_g, p):
    return _procesar_ecofibras(
        rutas, "san_joaquin", "Ecofibras San Joaquín", "ECOFIBRAS SAN JOAQUÍN", REGION_SAN_JOAQUIN,
        destino_san_joaquin, transportistas, homolog_c, homolog_g, p,
    )


def procesar_irar_los_angeles(rutas, transportistas, homolog_c, homolog_g, p):
    p("\n── I.R.A.R Los Ángeles — planta + BO (hoja INGRESOS) ──")

    df = leer_fuente(rutas["irar"], "INGRESOS", p, headers_fallback=(8, 9, 7, 0))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["FECHA", "Fecha"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    mov_col     = tomar_columna(df, ["MOVIMIENTO", "Movimiento"])
    destino_col = tomar_columna(df, ["DESTINO", "Destino", "DESTINO FINAL"])

    tmp = pd.DataFrame(index=df.index)
    tmp["Movimiento_final"] = mov_col.apply(lambda x: normalizar_movimiento(x, default="Ingreso"))
    tmp["DESTINO_COL"]      = destino_col

    p(f"  • Distribución Movimiento: {tmp['Movimiento_final'].value_counts(dropna=False).to_dict()}")

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["CLIENTE", "Cliente"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["RUT CLIENTE", "RUT"]).apply(limpiar_texto).values
    out["Gestor"]            = tomar_columna(df, ["TRANSPORTISTA", "GESTOR"], default="AMBIPAR ENVIRONMENT CHILE").apply(limpiar_texto).values
    out["Contrato"]          = tomar_columna(df, ["CONTRATO", "N° CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]         = tomar_columna(df, ["GENERADOR", "Generador"]).apply(limpiar_texto).values
    out["Transportista"]     = tomar_columna(df, ["TRANSPORTISTA", "Transportista"], default="AMBIPAR ENVIRONMENT CHILE").apply(limpiar_texto).values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Patente de Camión"] = tomar_columna(df, ["PATENTE", "PATENTE CAMIÓN", "Patente de Camión"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["TICKET DE PESAJE", "Ticket de pesaje"]).apply(limpiar_texto).values
    out["Peso neto"]         = tomar_columna(df, ["PESO NETO KG", "Peso neto", "PESO NETO"]).apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Movimiento"]        = tmp["Movimiento_final"].values
    out["Destino"]           = tmp.apply(destino_irar, axis=1).values
    out["Comuna Destino"]    = out["Destino"].apply(comuna_por_destino)
    out["TIPO"]              = tomar_columna(df, ["TIPO RESIDUO", "TIPO"]).apply(limpiar_texto).values
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_IRAR
    out["Destino inferido"]  = "No"

    out = excluir_traslados_a_planta_propia(out, "IRAR Los Ángeles", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ IRAR final: {len(out)} filas")
    return out


def procesar_bo_chillan(rutas, transportistas, homolog_c, homolog_g, p):
    p("\n── BO Chillán — SIEMPRE Traslado (hoja TRASLADOS BO) ──")

    df = leer_fuente(rutas["bo_chillan"], "TRASLADOS BO", p, headers_fallback=(0, 1, 7, 8))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["Fecha", "FECHA"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    out["Fecha"]             = df["_Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = tomar_columna(df, ["Cliente", "CLIENTE"]).apply(limpiar_texto).values
    out["RUT"]               = tomar_columna(df, ["RUT", "RUT CLIENTE"]).apply(limpiar_texto).values
    out["Gestor"]            = "AMBIPAR ENVIRONMENT CHILE"
    out["Contrato"]          = tomar_columna(df, ["Contrato", "N° CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]         = tomar_columna(df, ["Generador", "GENERADOR"]).apply(limpiar_texto).values
    out["Transportista"]     = "AMBIPAR ENVIRONMENT CHILE"
    out["Rut transportista"] = "96824110-9"
    out["Patente de Camión"] = tomar_columna(df, ["Patente de Camión", "PATENTE DE CAMIÓN", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"]  = tomar_columna(df, ["Ticket de pesaje", "TICKET DE PESAJE"]).apply(limpiar_texto).values
    out["Peso neto"]         = tomar_columna(df, ["Peso neto", "PESO NETO KG", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]            = tomar_columna(df, ["Unidad", "UNIDAD"], default="kg").apply(limpiar_texto).fillna("kg").values
    out["Movimiento"]        = "Traslado"
    out["Destino"]           = tomar_columna(df, ["Destino", "DESTINO"]).apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]    = tomar_columna(df, ["Comuna destino", "Comuna Destino", "COMUNA DESTINO"], default=None).apply(limpiar_texto).values
    out["Comuna Destino"]    = out["Comuna Destino"].fillna(out["Destino"].apply(comuna_por_destino))
    out["TIPO"]              = tomar_columna(df, ["TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_BO_CHILLAN
    out["Destino inferido"]  = "No"

    out = excluir_traslados_a_planta_propia(out, "BO Chillán", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ BO Chillán final: {len(out)} filas")
    return out


def procesar_temuco(rutas, homolog_c, homolog_g, p):
    p("\n── BO Temuco ──")

    df = leer_fuente(rutas["temuco"], "BBDD", p, headers_fallback=(7, 8, 0))
    p(f"  • Filas leídas: {len(df)}")

    fecha_col = tomar_columna(df, ["Fecha", "FECHA"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    out["Fecha"]            = df["_Fecha"].values
    out["Mes"]              = out["Fecha"].apply(nombre_mes)
    out["Cliente"]          = tomar_columna(df, ["Cliente", "CLIENTE"]).apply(limpiar_texto).values
    out["RUT"]              = tomar_columna(df, ["RUT", "RUT CLIENTE"]).apply(limpiar_texto).values
    out["Gestor"]           = "AMBIPAR ENVIRONMENT CHILE"
    out["Contrato"]         = tomar_columna(df, ["Contrato", "N° CONTRATO", "Nº CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]        = tomar_columna(df, ["Generador", "GENERADOR"]).apply(limpiar_texto).values
    out["Transportista"]    = "AMBIPAR ENVIRONMENT CHILE"
    out["Rut transportista"]= "96824110-9"
    out["Patente de Camión"]= tomar_columna(df, ["Patente de Camión", "PATENTE DE CAMIÓN", "PATENTE", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"] = tomar_columna(df, ["Ticket de pesaje", "TICKET DE PESAJE"]).apply(limpiar_texto).values
    out["Peso neto"]        = tomar_columna(df, ["Peso neto", "PESO NETO KG", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]           = tomar_columna(df, ["Unidad", "UNIDAD"], default="kg").apply(limpiar_texto).fillna("kg").values
    out["Destino"]          = tomar_columna(df, ["Destino", "DESTINO"]).apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]   = tomar_columna(df, ["Comuna destino", "Comuna Destino", "COMUNA DESTINO"], default=None).apply(limpiar_texto).values
    out["Comuna Destino"]   = out["Comuna Destino"].fillna(out["Destino"].apply(comuna_por_destino))
    out["TIPO"]             = tomar_columna(df, ["TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento"]       = "Traslado"
    out["Región"]           = REGION_TEMUCO
    out["Destino inferido"] = "No"

    out = excluir_traslados_a_planta_propia(out, "BO Temuco", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ Temuco final: {len(out)} filas")
    return out


def procesar_proveedores(rutas, transportistas, homolog_c, homolog_g, p):
    p("\n── Proveedores externos (Zona Sur) ──")

    if rutas.get("proveedores") is None or not Path(rutas["proveedores"]).exists():
        p("  ⚠ Archivo de proveedores no encontrado — fuente omitida")
        return base_final(0)

    df = leer_tabla_nombrada(rutas["proveedores"], "Tabla2", p)
    p(f"  • Filas leídas (Tabla2): {len(df)}")

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    n_antes = len(df)
    df = df[df["Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    out["Fecha"]             = df["Fecha"].values
    out["Mes"]               = out["Fecha"].apply(nombre_mes)
    out["Cliente"]           = df["Cliente"].apply(limpiar_texto).values
    out["RUT"]               = df["RUT CLIENTE"].apply(limpiar_texto).values
    gestor = df["Gestor"].apply(normalizar_gestor_ambipar)
    n_norm = int((gestor == GESTOR_AMBIPAR).sum())
    if n_norm:
        p(f"  • Gestor normalizado a {GESTOR_AMBIPAR}: {n_norm} filas")
    out["Gestor"]            = gestor.values
    out["Contrato"]          = None
    out["Generador"]         = df["Generador"].apply(limpiar_texto).values
    out["Transportista"]     = gestor.values
    out["Rut transportista"] = out["Transportista"].map(transportistas)
    out["Rut transportista"] = out["Rut transportista"].fillna(
        out["Transportista"].apply(lambda x: RUT_AMBIPAR if x == GESTOR_AMBIPAR else None)
    )
    out["Patente de Camión"] = df["Patente"].apply(limpiar_texto).values
    out["Ticket de pesaje"]  = df["N° Guía"].apply(limpiar_texto).values
    out["Peso neto"]         = df["Cantidad (Kg)"].apply(limpiar_numero).values
    out["Unidad"]            = "kg"
    out["Destino"]           = df["Destinatario"].apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]    = df["Comuna Destino"].apply(limpiar_texto).values
    out["TIPO"]              = df["TIPO"].apply(limpiar_texto).values
    out["Movimiento"]        = "Traslado externo"
    out["Movimiento interempresa"] = "No"
    out["Región"]            = REGION_PROVEEDORES
    out["Destino inferido"]  = "No"

    sin_rut = out[out["Rut transportista"].isna()]["Transportista"].dropna().unique()
    if len(sin_rut) > 0:
        p(f"  ⚠ Gestores sin RUT en el maestro de transportistas: {len(sin_rut)}")
        for g in sin_rut[:5]:
            p(f"     → {g!r}")

    p(f"  • Destinatarios: {out['Destino'].value_counts(dropna=False).to_dict()}")

    out = excluir_traslados_a_planta_propia(out, "Proveedores externos", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ Proveedores externos final: {len(out)} filas")
    return out


def procesar_bo_chiloe(rutas, homolog_c, homolog_g, p):
    p("\n── BO Chiloé ──")

    if rutas["bo_chiloe"] is None or not Path(rutas["bo_chiloe"]).exists():
        p("  ⚠ Archivo BO Chiloé no encontrado — fuente omitida")
        return base_final(0)

    df = leer_fuente(rutas["bo_chiloe"], "BBDD", p, headers_fallback=(6, 7, 0))
    p(f"  • Filas leídas: {len(df)}")

    col_cli = buscar_columna(df, ["Cliente", "CLIENTE"], obligatorio=True)
    mask_mowi = df[col_cli].apply(es_cliente_mowi)
    p(f"  • Exclusión MOWI: {int(mask_mowi.sum())} filas excluidas, "
      f"{int((~mask_mowi).sum())} conservadas")
    df = df.loc[~mask_mowi].copy().reset_index(drop=True)

    fecha_col = tomar_columna(df, ["Fecha", "FECHA"])
    df["_Fecha"] = pd.to_datetime(fecha_col, errors="coerce")

    n_antes = len(df)
    df = df[df["_Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)].copy()
    p(f"  • Filtro Sur (>= {FECHA_DESDE_SUR}): {len(df)} filas (excluidas: {n_antes - len(df)})")

    out = base_final(len(df))

    out["Fecha"]            = df["_Fecha"].values
    out["Mes"]              = out["Fecha"].apply(nombre_mes)
    out["Cliente"]          = tomar_columna(df, ["Cliente", "CLIENTE"]).apply(limpiar_texto).values
    out["RUT"]              = tomar_columna(df, ["RUT", "RUT CLIENTE"]).apply(limpiar_texto).values
    out["Gestor"]           = "AMBIPAR ENVIRONMENT CHILE"
    out["Contrato"]         = tomar_columna(df, ["Contrato", "N° CONTRATO", "Nº CONTRATO"]).apply(limpiar_texto).values
    out["Generador"]        = tomar_columna(df, ["Generador", "GENERADOR"]).apply(limpiar_texto).values
    out["Transportista"]    = "AMBIPAR ENVIRONMENT CHILE"
    out["Rut transportista"]= "96824110-9"
    out["Patente de Camión"]= tomar_columna(df, ["Patente de Camión", "PATENTE DE CAMIÓN", "PATENTE", "Patente"]).apply(limpiar_texto).values
    out["Ticket de pesaje"] = tomar_columna(df, ["Ticket de pesaje", "TICKET DE PESAJE"]).apply(limpiar_texto).values
    out["Peso neto"]        = tomar_columna(df, ["Peso neto [kg]", "Peso neto", "PESO NETO KG", "Peso neto (kg)"]).apply(limpiar_numero).values
    out["Unidad"]           = tomar_columna(df, ["Unidad", "UNIDAD"], default="kg").apply(limpiar_texto).fillna("kg").values
    out["Destino"]          = tomar_columna(df, ["Destino", "DESTINO"]).apply(lambda x: normalizar_destino(x)).values
    out["Comuna Destino"]   = tomar_columna(df, ["Comuna destino", "Comuna Destino", "COMUNA DESTINO"], default=None).apply(limpiar_texto).values
    out["Comuna Destino"]   = out["Comuna Destino"].fillna(out["Destino"].apply(comuna_por_destino))
    out["TIPO"]             = tomar_columna(df, ["TIPO", "TIPO RESIDUO"]).apply(limpiar_texto).values
    out["Movimiento"]       = "Traslado"
    out["Región"]           = REGION_BO_CHILOE
    out["Destino inferido"] = "No"

    out = excluir_traslados_a_planta_propia(out, "BO Chiloé", p)

    out = aplicar_homologacion_cliente(out, "Cliente", "RUT", homolog_c)
    out = aplicar_homologacion_generador(out, "Generador", homolog_g)

    p(f"  ✓ BO Chiloé final: {len(out)} filas")
    return out


# ══════════════════════════════════════════════════════════════════
# JOINS
# ══════════════════════════════════════════════════════════════════
def unir_sinader(df, materiales, p):
    p("\n── Join con SINADER (TIPO → CÓDIGOS SINADER) ──")

    df["CÓDIGOS SINADER"] = df["TIPO"].astype(str).str.strip().map(materiales)

    sin = df[df["CÓDIGOS SINADER"].isna() & df["TIPO"].notna()][["__origen__", "TIPO"]]

    if len(sin) > 0:
        p(f"  ⚠ {len(sin)} filas sin código SINADER")
        for _, r in sin.drop_duplicates().head(20).iterrows():
            p(f"     [{r['__origen__']}] {r['TIPO']}")
    else:
        p("  ✓ Todos los TIPO tienen código SINADER")

    return df


def unir_destinatarios(df, df_dest, p):
    p("\n── Join con DESTINATARIOS SUR (Destino + Comuna + LER + Región) ──")

    df = df.copy()
    dj = df_dest.copy()

    if "Comuna Destino" in df.columns:
        lookup_comuna = (
            dj[["NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", "REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"]]
            .drop_duplicates()
            .copy()
        )
        lookup_comuna["_k_dest"]   = lookup_comuna["NOMBRE DE FANTASÍA"].apply(normalizar_texto)
        lookup_comuna["_k_region"] = lookup_comuna["REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"].apply(normalizar_region_para_cruce)

        dup = lookup_comuna.duplicated(subset=["_k_dest", "_k_region"], keep=False)
        if dup.any():
            amb = (
                lookup_comuna.loc[dup, ["NOMBRE DE FANTASÍA", "REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL", "COMUNA DE ESTABLECIMIENTO"]]
                .drop_duplicates()
                .values.tolist()
            )
            p(f"  ⚠ Destinos con más de una comuna en la misma región (se usa la primera): {amb}")

        lookup_todo = lookup_comuna.copy()

        lookup_comuna = lookup_comuna.drop_duplicates(subset=["_k_dest", "_k_region"], keep="first")

        mask = df["Comuna Destino"].isna() & df["Destino"].notna()
        if mask.any():
            tmp_base = df.loc[mask, ["Destino", "Región"]].copy()
            tmp_base["_k_dest"]   = tmp_base["Destino"].apply(normalizar_texto)
            tmp_base["_k_region"] = tmp_base["Región"].apply(normalizar_region_para_cruce)

            tmp = tmp_base.merge(
                lookup_comuna[["_k_dest", "_k_region", "COMUNA DE ESTABLECIMIENTO"]],
                on=["_k_dest", "_k_region"],
                how="left",
            )

            falta = tmp["COMUNA DE ESTABLECIMIENTO"].isna()
            if falta.any():
                comunas_por_dest = (
                    lookup_todo[["_k_dest", "COMUNA DE ESTABLECIMIENTO"]]
                    .drop_duplicates()
                    .groupby("_k_dest")["COMUNA DE ESTABLECIMIENTO"]
                    .agg(list)
                )
                unicos = comunas_por_dest[comunas_por_dest.apply(len) == 1]
                ambiguos = comunas_por_dest[comunas_por_dest.apply(len) > 1]

                lookup_solo_dest = pd.DataFrame({
                    "_k_dest": unicos.index,
                    "_comuna_fb": [v[0] for v in unicos.values],
                })

                tmp = tmp.merge(lookup_solo_dest, on="_k_dest", how="left")
                tmp["COMUNA DE ESTABLECIMIENTO"] = tmp["COMUNA DE ESTABLECIMIENTO"].fillna(tmp["_comuna_fb"])

                n_fb = int(tmp.loc[falta, "COMUNA DE ESTABLECIMIENTO"].notna().sum())
                if n_fb:
                    p(f"  • Comuna rellenada por nombre de destino (destinatario en otra región): {n_fb} filas")

                pendientes = set(tmp.loc[tmp["COMUNA DE ESTABLECIMIENTO"].isna(), "_k_dest"])
                choque = sorted(pendientes & set(ambiguos.index))
                if choque:
                    p(f"  ⚠ Comuna NO rellenada: estos destinos tienen varias comunas en el maestro "
                      f"y no se puede deducir cuál corresponde. Completar la Comuna Destino en el archivo fuente:")
                    for k in choque[:10]:
                        p(f"     · {k} → comunas posibles: {ambiguos[k]}")

            if len(tmp) != int(mask.sum()):
                p(f"  ⚠ Merge de comuna cambió el largo ({int(mask.sum())} → {len(tmp)}); "
                  f"se omite el relleno automático de Comuna Destino. REVISAR maestro de destinatarios.")
            else:
                df.loc[mask, "Comuna Destino"] = tmp["COMUNA DE ESTABLECIMIENTO"].values
                n_rell = int(tmp["COMUNA DE ESTABLECIMIENTO"].notna().sum())
                p(f"  • Comuna Destino rellenada desde maestro: {n_rell}/{int(mask.sum())} filas")

    df["_k_dest"]   = df["Destino"].apply(normalizar_texto)
    df["_k_comuna"] = df["Comuna Destino"].apply(normalizar_texto)
    df["_k_ler"]    = df["CÓDIGOS SINADER"].apply(normalizar_ler)
    df["_k_region"] = df["Región"].apply(normalizar_region_para_cruce)

    dj["_k_dest"]   = dj["NOMBRE DE FANTASÍA"].apply(normalizar_texto)
    dj["_k_comuna"] = dj["COMUNA DE ESTABLECIMIENTO"].apply(normalizar_texto)
    dj["_k_region"] = dj["REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"].apply(normalizar_region_para_cruce)

    dj["_ler_lista"] = dj["CÓDIGOS LER"].astype(str).apply(
        lambda s: [t for t in re.split(r"[;,\n\r/|]+", s) if t.strip()] or [s]
    )
    n_dj_antes = len(dj)
    dj = dj.explode("_ler_lista")
    dj["_k_ler"] = dj["_ler_lista"].apply(normalizar_ler)
    if len(dj) != n_dj_antes:
        p(f"  • Maestro expandido por multi-LER: {n_dj_antes} → {len(dj)} filas")

    cols_merge = [
        "_k_dest", "_k_comuna", "_k_ler", "_k_region",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER",
    ]

    dj_merge = dj[cols_merge].drop_duplicates(subset=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], keep="first")

    n_antes = len(df)

    # Pasada 1: cruce con región
    df = df.merge(dj_merge, on=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], how="left")
    df = df.sort_values("__idx__").drop_duplicates(subset=["__idx__"], keep="first")

    n_match_p1 = df["RUT DESTINATARIO"].notna().sum()
    p(f"  Pasada 1 (con región): {n_match_p1}/{n_antes} matchearon")

    # Pasada 2: sin región
    sin_match = df["RUT DESTINATARIO"].isna()
    if sin_match.any():
        dj_sin_region = dj[
            ["_k_dest", "_k_comuna", "_k_ler", "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER"]
        ].drop_duplicates(subset=["_k_dest", "_k_comuna", "_k_ler"], keep="first")

        tmp_sin = df.loc[sin_match, ["__idx__", "_k_dest", "_k_comuna", "_k_ler"]].merge(
            dj_sin_region, on=["_k_dest", "_k_comuna", "_k_ler"], how="left"
        )

        matched_p2 = tmp_sin["RUT DESTINATARIO"].notna()
        n_match_p2 = matched_p2.sum()
        p(f"  Pasada 2 (sin región): {n_match_p2} adicionales matchearon")

        if n_match_p2 > 0:
            idx_p2 = tmp_sin.loc[matched_p2, "__idx__"].values
            for col in ["RUT DESTINATARIO", "CÓDIGO ESTABLECIMIENTO SINADER", "CÓDIGO DE TRATAMIENTO SINADER"]:
                df.loc[df["__idx__"].isin(idx_p2), col] = (
                    tmp_sin.loc[matched_p2].set_index("__idx__")[col].reindex(idx_p2).values
                )

    # Pasada 3: solo destino + LER
    sin_match3 = df["RUT DESTINATARIO"].isna()
    if sin_match3.any():
        ruts_por_par = (
            dj[["_k_dest", "_k_ler", "RUT DESTINATARIO"]]
            .dropna(subset=["RUT DESTINATARIO"])
            .drop_duplicates()
            .groupby(["_k_dest", "_k_ler"])["RUT DESTINATARIO"]
            .nunique()
        )
        pares_unicos = set(ruts_por_par[ruts_por_par == 1].index)
        pares_ambiguos = sorted(ruts_por_par[ruts_por_par > 1].index)[:10]

        dj_dest_ler = dj[
            ["_k_dest", "_k_ler", "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO", "CÓDIGO DE TRATAMIENTO SINADER"]
        ].copy()
        dj_dest_ler = dj_dest_ler[
            dj_dest_ler.set_index(["_k_dest", "_k_ler"]).index.isin(pares_unicos)
        ].drop_duplicates(subset=["_k_dest", "_k_ler"], keep="first")

        tmp3 = df.loc[sin_match3, ["__idx__", "_k_dest", "_k_ler"]].merge(
            dj_dest_ler, on=["_k_dest", "_k_ler"], how="left"
        )

        matched_p3 = tmp3["RUT DESTINATARIO"].notna()
        n_match_p3 = int(matched_p3.sum())
        p(f"  Pasada 3 (destino + LER, solo casos inequívocos): {n_match_p3} adicionales matchearon")

        if pares_ambiguos:
            p(f"  ⚠ Pares destino+LER con más de un RUT destinatario (NO se cruzaron): {pares_ambiguos}")

        if n_match_p3 > 0:
            idx_p3 = tmp3.loc[matched_p3, "__idx__"].values
            for col in ["RUT DESTINATARIO", "CÓDIGO ESTABLECIMIENTO SINADER", "CÓDIGO DE TRATAMIENTO SINADER"]:
                df.loc[df["__idx__"].isin(idx_p3), col] = (
                    tmp3.loc[matched_p3].set_index("__idx__")[col].reindex(idx_p3).values
                )

    df = df.drop(columns=["_k_dest", "_k_comuna", "_k_ler", "_k_region"], errors="ignore")

    p(f"  {'✓' if len(df) == n_antes else '⚠'} JOIN TOTAL: {n_antes} → {len(df)} filas")

    sin = df[df["RUT DESTINATARIO"].isna()][
        ["__origen__", "Destino", "Comuna Destino", "Región", "CÓDIGOS SINADER"]
    ].copy()

    if len(sin) > 0:
        agrupado = (
            sin.groupby(["Destino", "Comuna Destino", "Región", "CÓDIGOS SINADER"], dropna=False)["__origen__"]
            .agg(lambda x: ", ".join(sorted(set(map(str, x)))))
            .reset_index()
            .rename(columns={"__origen__": "ORIGENES"})
        )

        p(f"  ⚠ {len(agrupado)} combinaciones sin match en Destinatarios:")
        for _, r in agrupado.head(30).iterrows():
            p(
                f"     [{r['ORIGENES']}] {r['Destino']} | {r['Comuna Destino']} | "
                f"{r['Región']} | {r['CÓDIGOS SINADER']}"
            )

        p("  ── Comparación contra el maestro de destinatarios ──")
        for d in sorted(set(sin["Destino"].dropna().astype(str)))[:15]:
            k = normalizar_texto(d)
            cand = dj[dj["_k_dest"] == k]
            if cand.empty:
                p(f"     ✗ '{d}': NO existe en el maestro. Revisar NOMBRE DE FANTASÍA "
                  f"(o que la región del destinatario no sea RM, que se excluye al cargar).")
                continue
            comunas  = sorted(set(cand["COMUNA DE ESTABLECIMIENTO"].dropna().astype(str)))
            regiones = sorted(set(cand["REGIÓN DE ESTABLECIMIENTO NORMALIZADA_ORIGINAL"].dropna().astype(str)))
            lers     = sorted(set(cand["_k_ler"].dropna().astype(str)))
            lers_mov = sorted(set(sin.loc[sin["Destino"].astype(str) == d, "CÓDIGOS SINADER"]
                                  .dropna().astype(str).apply(normalizar_ler)))
            p(f"     ℹ '{d}' en maestro → comuna: {comunas} | región: {regiones}")
            p(f"        LER maestro: {lers[:12]}")
            p(f"        LER movimiento: {lers_mov}")
            faltan = [x for x in lers_mov if x and x not in lers]
            if faltan:
                p(f"        ✗ CAUSA: el código {faltan} no está autorizado para este destinatario "
                  f"en el maestro. Agregarlo a CÓDIGOS LER o revisar el TIPO del movimiento.")

        pm = agrupado[agrupado["Destino"].astype(str).str.upper().str.contains("TRAP|SAN JOAQU", na=False, regex=True)]
        if len(pm) > 0:
            p("  ⚠ Nota: si ECOFIBRAS TRAPÉN o ECOFIBRAS SAN JOAQUÍN no cruzan, revisa que existan")
            p(f"     en la hoja '{HOJA_DESTINATARIOS}' con columna CÓDIGOS LER.")

    else:
        p("  ✓ Todos los destinos encontraron match en Destinatarios")

    return df


def limpiar_filas_invalidas(df, p):
    claves = ["Fecha", "Cliente", "Destino", "TIPO"]
    temp = df[claves].copy()

    for c in claves:
        temp[c] = temp[c].astype(str).str.strip().str.lower()

    vacia = temp.isin(["", "nan", "none", "nat"]).all(axis=1)
    n = int(vacia.sum())

    if n > 0:
        p(f"  ⚠ Filas inválidas eliminadas: {n}")

    return df.loc[~vacia].copy()


# ══════════════════════════════════════════════════════════════════
# PIPELINE
# ══════════════════════════════════════════════════════════════════
def consolidar(rutas, ruta_prueba=None, ruta_log=None, modo_reset=False,
               modo_prueba=False, mostrar=True):
    """Ejecuta la consolidación completa de Sur.

    Ver docstring de la versión RM para el detalle de parámetros y de lo que
    devuelve — el patrón es idéntico.
    """
    p = Registro(mostrar=mostrar, ruta_log=ruta_log)

    verificar_rutas(rutas)
    if modo_prueba and not ruta_prueba:
        raise ValueError("modo_prueba=True requiere indicar ruta_prueba.")

    p("\n" + "═" * 72)
    p("  INICIO CONSOLIDACIÓN ZONA SUR" + (" — MODO PRUEBA" if modo_prueba else ""))
    p(f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    p(f"  Procesando registros desde: {FECHA_DESDE_SUR}")
    p("═" * 72)

    homolog_c, homolog_g = cargar_homologacion(rutas["homologacion"], p)
    transportistas, materiales, df_dest = cargar_lookups(rutas, p)

    df_coronel       = procesar_ecofibras_coronel(rutas, transportistas, homolog_c, homolog_g, p)
    df_bo_trapen     = procesar_bo_trapen(rutas, transportistas, homolog_c, homolog_g, p)
    df_planta_trapen = procesar_ecofibras_planta_trapen(rutas, transportistas, homolog_c, homolog_g, p)
    df_san_joaquin   = procesar_ecofibras_san_joaquin(rutas, transportistas, homolog_c, homolog_g, p)
    df_irar          = procesar_irar_los_angeles(rutas, transportistas, homolog_c, homolog_g, p)
    df_bo_chillan    = procesar_bo_chillan(rutas, transportistas, homolog_c, homolog_g, p)
    df_temuco        = procesar_temuco(rutas, homolog_c, homolog_g, p)
    df_bo_chiloe     = procesar_bo_chiloe(rutas, homolog_c, homolog_g, p)
    df_proveedores   = procesar_proveedores(rutas, transportistas, homolog_c, homolog_g, p)

    fuentes = {
        "CORONEL":          df_coronel,
        "BO_TRAPEN":        df_bo_trapen,
        "PLANTA_TRAPEN":    df_planta_trapen,
        "SAN_JOAQUIN":      df_san_joaquin,
        "IRAR_LOS_ANGELES": df_irar,
        "BO_CHILLAN":       df_bo_chillan,
        "TEMUCO":           df_temuco,
        "BO_CHILOE":        df_bo_chiloe,
        "PROVEEDORES":      df_proveedores,
    }

    p("\n── Combinando fuentes Zona Sur ──")

    suma_fuentes = 0
    lista = []

    for nombre, dfi in fuentes.items():
        dfi = dfi.copy()
        dfi["__origen__"] = nombre
        lista.append(dfi)
        p(f"  {nombre:<20} {len(dfi):>6} filas")
        suma_fuentes += len(dfi)

    p(f"  {'─' * 30}")
    p(f"  Suma esperada:       {suma_fuentes:>6} filas")

    df = pd.concat(lista, ignore_index=True)

    # Filtro RESIMPLE global
    mask_res = df["Cliente"].apply(es_cliente_resimple)
    if mask_res.any():
        detalle = df.loc[mask_res].groupby("__origen__").size().to_dict()
        p(f"  • Filtro RESIMPLE global: excluidas {int(mask_res.sum())} filas")
        for origen, cnt in detalle.items():
            p(f"     → {origen}: {cnt} filas")
        df = df.loc[~mask_res].copy().reset_index(drop=True)
    else:
        p("  ✓ Filtro RESIMPLE global: 0 filas con cliente Resimple")

    df["__idx__"] = df.index

    if len(df) == suma_fuentes:
        p(f"  ✓ Combinado: {len(df)} filas (0 pérdidas)")
    else:
        p(f"  ⚠ Diferencia en combinación: esperadas {suma_fuentes}, obtenidas {len(df)}")

    df["Destino inferido"] = "No"

    if "Movimiento interempresa" not in df.columns:
        df["Movimiento interempresa"] = "No"
    df["Movimiento interempresa"] = df["Movimiento interempresa"].fillna("No")

    df = unir_sinader(df, materiales, p)
    df = unir_destinatarios(df, df_dest, p)

    df = df.drop(columns=["Unidad", "__origen__", "__idx__"], errors="ignore")
    df = df.rename(columns={"Peso neto": "Peso neto (kg)"})

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["Mes"]   = df["Fecha"].apply(nombre_mes)

    df = limpiar_filas_invalidas(df, p)

    for col in COLUMNAS_FINALES:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMNAS_FINALES]

    p(f"\n── DataFrame final SUR: {len(df)} filas × {len(df.columns)} columnas ──")

    if modo_prueba:
        ruta_prueba = Path(ruta_prueba)
        # Sello de fecha/hora para no pisar corridas anteriores (igual que RM).
        sello = datetime.now().strftime("%Y-%m-%d_%H%M")
        ruta_prueba = ruta_prueba.with_name(f"{ruta_prueba.stem}_{sello}{ruta_prueba.suffix}")
        p(f"\n── Modo PRUEBA — guardando {ruta_prueba.name} ──")
        guardar_excel(df, ruta_prueba, p)
        df_final = df
    else:
        ruta_destino = Path(rutas["destino_real"])

        if modo_reset:
            p("\n── Modo RESET — reemplazando BBDD_TRAZABILIDAD_SUR.xlsx ──")
            df_final = df.copy()
        else:
            p("\n── Modo ACUMULATIVO ──")

            if ruta_destino.exists():
                df_existente = leer_excel(ruta_destino, p, sheet_name=HOJA_DESTINO)
                p(f"  Trazabilidad SUR existente: {len(df_existente)} filas")
            else:
                df_existente = pd.DataFrame(columns=COLUMNAS_FINALES)
                p("  Archivo SUR final no existe — se creará nuevo")

            if not df_existente.empty:
                temp_exist = df_existente.copy()
                temp_new   = df.copy()

                for col in CLAVE_DEDUP:
                    temp_exist[col] = temp_exist[col].astype(str).str.strip() if col in temp_exist.columns else ""
                    temp_new[col]   = temp_new[col].astype(str).str.strip()   if col in temp_new.columns   else ""

                llave_exist = temp_exist[CLAVE_DEDUP].apply(tuple, axis=1)
                llave_nueva = temp_new[CLAVE_DEDUP].apply(tuple, axis=1)

                df_nuevas = df[~llave_nueva.isin(llave_exist)].copy()
            else:
                df_nuevas = df.copy()

            p(f"  Filas nuevas a agregar: {len(df_nuevas)}")
            p(f"  Duplicadas omitidas:    {len(df) - len(df_nuevas)}")

            df_final = pd.concat([df_existente, df_nuevas], ignore_index=True)

        for col in COLUMNAS_FINALES:
            if col not in df_final.columns:
                df_final[col] = None
        df_final = df_final[COLUMNAS_FINALES]

        p(f"  Total final SUR: {len(df_final)} filas")
        guardar_excel(df_final, ruta_destino, p)

    p("\n" + "═" * 72)
    p("  CONSOLIDACIÓN SUR COMPLETADA ✓")
    p("═" * 72 + "\n")

    return {
        "consolidado": df_final,
        "filas": len(df_final),
        "columnas": len(df_final.columns),
        "log": p.texto(),
    }


# ══════════════════════════════════════════════════════════════════
# USO DESDE LA TERMINAL
# ══════════════════════════════════════════════════════════════════
def _main():
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas, ConfiguracionFaltante

    modo_reset = "--reset" in sys.argv
    modo_prueba = "--prueba" in sys.argv

    try:
        rutas = rutas_desde_config()
        c = carpetas()
    except ConfiguracionFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    ruta_prueba = c["sur"] / "PRUEBA_TRAZABILIDAD_SUR.xlsx"
    ruta_log = c["sur"] / "log_consolidacion_sur.txt"

    try:
        consolidar(
            rutas,
            ruta_prueba=ruta_prueba,
            ruta_log=ruta_log,
            modo_reset=modo_reset,
            modo_prueba=modo_prueba,
        )
    except (ArchivoFaltante, PermissionError, ValueError) as e:
        print(f"\n✗ ERROR CRÍTICO: {e}\n")
        sys.exit(2)


if __name__ == "__main__":
    _main()
