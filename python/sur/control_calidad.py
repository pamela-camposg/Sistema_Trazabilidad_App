"""
control_calidad.py — Control de calidad Zona Sur (versión refactorizada)
Ambipar Group — Proyecto TRZ-APP-001, Etapa 1

Mismo patrón que src/rm/control_calidad.py: la lógica es la misma que en
01_control_calidad_sur.py, línea por línea. Cambia solo cómo se reciben las
rutas (parámetro, no rutas fijas de OneDrive) y que la función principal
devuelve los resultados además de escribir el Excel.

Reglas clave de Sur (documentadas en el script original, sin cambios):
    - Zona Sur NO infiere destinos por histórico.
    - I.R.A.R Los Ángeles usa SOLO la hoja TRASLADOS BO.
    - Coronel / San Joaquín:
        Movimiento = Ingreso  → destino = planta
        Movimiento = Traslado → destino = columna DESTINO
    - Ecofibras Planta Trapén: NO maneja Movimiento → TODO es Ingreso.
    - BO Trapén: SIEMPRE es Traslado.
    - Temuco: solo traslados.
    - IRAR: hoja TRASLADOS BO → destino = columna Destino.
    - BO Chiloé: hoja BBDD, solo traslados; se EXCLUYE el cliente MOWI.

Se puede usar de dos formas:

    Desde otro programa:
        from control_calidad import controlar, rutas_desde_carpeta
        rutas = rutas_desde_carpeta("C:/carpeta/con/los/excel")
        resultado = controlar(rutas)

    Desde la terminal, igual que antes:
        python control_calidad.py
"""

import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook


# ══════════════════════════════════════════════════════════════════
# CONSTANTES DE LA ZONA SUR — reglas de negocio, no rutas
# ══════════════════════════════════════════════════════════════════
FECHA_DESDE_SUR = "2026-04-01"
HOJA_DESTINO = "TRAZABILIDAD"

MAX_REINTENTOS = 5
ESPERA_REINTENTO = 30

# Nombre esperado de cada archivo fuente. La clave es el nombre interno; el
# valor son los nombres posibles del archivo (varían en OneDrive).
NOMBRES_ARCHIVO = {
    "coronel":        ["BBDD ECOFIBRAS CORONEL.xlsx"],
    "bo_trapen":       ["BBDD BO TRAPEN.xlsx"],
    "planta_trapen":   ["BBDD ECOFIBRAS PLANTA TRAPEN .xlsx", "BBDD ECOFIBRAS PLANTA TRAPEN.xlsx"],
    "san_joaquin":     ["BBDD ECOFIBRAS SAN JOAQUIN.xlsx"],
    "irar":            ["BBDD IRAR LOS ÁNGELES.xlsx", "BBDD I.R.A.R LOS ÁNGELES.xlsx"],
    "bo_chillan":      ["BBDD BO CHILLÁN.xlsx"],
    "temuco":          ["BBDD TEMUCO.xlsx"],
    "bo_chiloe":       ["BBDD BO CHILOÉ.xlsx", "BBDD BO CHILOE.xlsx"],
    "homologacion":    ["HOMOLOGACION.xlsx"],
    "destinatarios":   ["BBDD_DESTINATARIO.xlsx", "BBDD DESTINATARIO.xlsx"],
    "sinader":         ["Clasificación_Residuos SINADER.xlsx", "Clasificacion_Residuos SINADER.xlsx"],
    "transportistas":  ["Transportistas.xlsx"],
}

# bo_chiloe es opcional en el script original (se omite si no existe el
# archivo). El resto es obligatorio.
OPCIONALES = {"bo_chiloe"}


class ArchivoFaltante(Exception):
    """Se lanza cuando falta algún archivo fuente obligatorio."""


# ══════════════════════════════════════════════════════════════════
# REGISTRO DE MENSAJES
# ══════════════════════════════════════════════════════════════════
class Registro:
    def __init__(self, mostrar=True):
        self.lineas = []
        self.mostrar = mostrar

    def __call__(self, msg=""):
        self.lineas.append(str(msg))
        if self.mostrar:
            print(msg)

    def texto(self):
        return "\n".join(self.lineas)


# ══════════════════════════════════════════════════════════════════
# CÓMO ENCONTRAR LOS ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def _clave(nombre):
    """Normaliza un nombre de archivo para comparar tolerando tildes/espacios."""
    n = unicodedata.normalize("NFKD", str(nombre))
    n = n.encode("ascii", "ignore").decode("ascii")
    n = n.upper()
    for c in " ._-()":
        n = n.replace(c, "")
    return n


def rutas_desde_config(base=None):
    """Arma las rutas usando la estructura de carpetas de siempre en OneDrive."""
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from config import carpetas

    c = carpetas(base)
    zona = c["zona_sur"]
    info = c["info_base"]

    rutas = {
        "coronel":        zona / "BBDD ECOFIBRAS CORONEL.xlsx",
        "bo_trapen":       zona / "BBDD BO TRAPEN.xlsx",
        "planta_trapen":   zona / "BBDD ECOFIBRAS PLANTA TRAPEN .xlsx",
        "san_joaquin":     zona / "BBDD ECOFIBRAS SAN JOAQUIN.xlsx",
        "irar":            zona / "BBDD IRAR LOS ÁNGELES.xlsx",
        "bo_chillan":      zona / "BBDD BO CHILLÁN.xlsx",
        "temuco":          zona / "BBDD TEMUCO.xlsx",
        "bo_chiloe":       zona / "BBDD BO CHILOÉ.xlsx",
        "homologacion":    info / "HOMOLOGACION.xlsx",
        "destinatarios":   info / "BBDD_DESTINATARIO.xlsx",
        "sinader":         info / "Clasificación_Residuos SINADER.xlsx",
        "transportistas":  info / "Transportistas.xlsx",
    }
    # irar tiene una variante alternativa de nombre; usar la que exista
    alt_irar = zona / "BBDD I.R.A.R LOS ÁNGELES.xlsx"
    if not rutas["irar"].exists() and alt_irar.exists():
        rutas["irar"] = alt_irar
    alt_dest = info / "BBDD DESTINATARIO.xlsx"
    if not rutas["destinatarios"].exists() and alt_dest.exists():
        rutas["destinatarios"] = alt_dest

    return rutas


def rutas_desde_carpeta(carpeta):
    """Arma las rutas buscando todos los archivos dentro de una sola carpeta.

    Es la forma que va a usar la aplicación cuando la operadora suba los
    archivos. bo_chiloe es opcional; el resto es obligatorio.
    """
    carpeta = Path(carpeta)
    if not carpeta.exists():
        raise ArchivoFaltante(f"La carpeta no existe:\n  {carpeta}")

    presentes = {_clave(f.name): f for f in carpeta.iterdir() if f.is_file()}

    rutas = {}
    faltantes = []
    for interno, posibles in NOMBRES_ARCHIVO.items():
        encontrado = None
        for nombre in posibles:
            encontrado = presentes.get(_clave(nombre))
            if encontrado:
                break
        if encontrado:
            rutas[interno] = encontrado
        elif interno not in OPCIONALES:
            faltantes.append(posibles[0])

    if faltantes:
        hay = "\n".join(f"    · {f.name}" for f in sorted(presentes.values()))
        falta = "\n".join(f"    · {f}" for f in faltantes)
        raise ArchivoFaltante(
            f"Faltan archivos en la carpeta:\n{falta}\n\n"
            f"  Archivos encontrados en {carpeta.name}:\n{hay if hay else '    (ninguno)'}"
        )

    return rutas


def verificar_rutas(rutas):
    obligatorias = [k for k in NOMBRES_ARCHIVO if k not in OPCIONALES]
    faltan = [f"{k}: {rutas[k]}" for k in obligatorias if k not in rutas or not Path(rutas[k]).exists()]
    if faltan:
        detalle = "\n".join(f"    · {f}" for f in faltan)
        raise ArchivoFaltante(f"No se encontraron estos archivos:\n{detalle}")


# ══════════════════════════════════════════════════════════════════
# FUNCIONES DE TEXTO — idénticas a 01_control_calidad_sur.py
# ══════════════════════════════════════════════════════════════════
def nombre_mes(fecha):
    meses = {
        1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
        5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
        9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
    }
    try:
        return meses[pd.Timestamp(fecha).month]
    except Exception:
        return ""


def limpiar_texto(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s.lower() in {"", "nan", "none", "nat", "#n/a", "n/a"}:
        return None
    return re.sub(r"\s+", " ", s).strip()


def normalizar_texto(x):
    s = limpiar_texto(x)
    if not s:
        return ""
    s = s.upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = s.replace(".", "")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalizar_nombre_para_buscar(x):
    return normalizar_texto(x)


def normalizar_movimiento(x, default=None):
    s = limpiar_texto(x)
    if not s:
        return default
    n = normalizar_texto(s)
    if "INGRES" in n:
        return "Ingreso"
    if "TRASL" in n:
        return "Traslado"
    if "SALIDA" in n:
        return "Salida"
    return s.strip()


def buscar_columna(df, opciones, obligatorio=False):
    """Busca una columna por nombre, tolerando mayúsculas, acentos y espacios."""
    mapa = {normalizar_texto(c): c for c in df.columns}
    for op in opciones:
        key = normalizar_texto(op)
        if key in mapa:
            return mapa[key]
    if obligatorio:
        raise ValueError(f"No encontré ninguna columna entre: {opciones}. Columnas disponibles: {list(df.columns)}")
    return None


def tomar_columna(df, opciones, default=None, obligatorio=False):
    col = buscar_columna(df, opciones, obligatorio=obligatorio)
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]


# ══════════════════════════════════════════════════════════════════
# LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════
def leer_excel(path, p, **kwargs):
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            return pd.read_excel(path, **kwargs)
        except PermissionError:
            p(f"  ⚠ Bloqueado: {Path(path).name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)
    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def leer_fuente(path, hoja, p, headers_fallback=(0, 7, 8, 9)):
    """Lee una fuente por hoja: primero intenta tabla Excel, luego headers alternativos."""
    path = Path(path)

    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            wb = load_workbook(path, data_only=True, read_only=False)

            if hoja not in wb.sheetnames:
                wb.close()
                raise ValueError(f"No existe la hoja '{hoja}' en {path.name}. Hojas: {wb.sheetnames}")

            ws = wb[hoja]

            if ws.tables:
                tabla = list(ws.tables.values())[0]
                filas = [[c.value for c in fila] for fila in ws[tabla.ref]]
                headers = filas[0]
                df = pd.DataFrame(filas[1:], columns=headers)
                wb.close()
                df.columns = df.columns.astype(str).str.strip()
                return df.dropna(how="all").reset_index(drop=True)

            wb.close()

            mejor = None
            mejor_score = -1
            claves = [
                "FECHA", "CLIENTE", "RUT", "GENERADOR", "TRANSPORTISTA",
                "DESTINO", "TIPO", "MOVIMIENTO", "PESO", "TICKET", "PATENTE",
            ]

            for header in headers_fallback:
                try:
                    cand = pd.read_excel(path, sheet_name=hoja, header=header)
                    cand.columns = cand.columns.astype(str).str.strip()
                    cand = cand.dropna(how="all").reset_index(drop=True)

                    score = 0
                    cols_norm = [normalizar_texto(c) for c in cand.columns]
                    for clave in claves:
                        if any(clave in c for c in cols_norm):
                            score += 1

                    if score > mejor_score:
                        mejor = cand
                        mejor_score = score
                except Exception:
                    pass

            if mejor is None:
                raise ValueError(f"No se pudo leer {path.name} / hoja {hoja}")

            return mejor

        except PermissionError:
            p(f"  ⚠ Bloqueado: {path.name} — intento {intento}/{MAX_REINTENTOS}")
            if intento < MAX_REINTENTOS:
                time.sleep(ESPERA_REINTENTO)

    raise PermissionError(f"No se pudo abrir {path}. Cierra el archivo en Excel.")


def cargar_homologacion(ruta_homologacion, p):
    homolog_clientes = {}
    homolog_generadores = {}

    ruta = Path(ruta_homologacion)
    p("\n── Cargando HOMOLOGACION desde OneDrive ──")
    p(f"  {ruta}")

    if not ruta.exists():
        p("  ⚠ No se encontró HOMOLOGACION.xlsx — se continúa sin correcciones")
        return homolog_clientes, homolog_generadores

    try:
        df_cli = leer_excel(ruta, p, sheet_name="clientes")
        df_cli = df_cli.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_cli.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            rut_correcto = limpiar_texto(row.get("RUT_CORRECTO"))
            if variante and nombre_correcto:
                homolog_clientes[variante] = {"nombre": nombre_correcto, "rut": rut_correcto}
        p(f"  Clientes homologados: {len(homolog_clientes)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja clientes: {e}")

    try:
        df_gen = leer_excel(ruta, p, sheet_name="generadores")
        df_gen = df_gen.dropna(subset=["NOMBRE_VARIANTE"])
        for _, row in df_gen.iterrows():
            variante = normalizar_nombre_para_buscar(row.get("NOMBRE_VARIANTE"))
            nombre_correcto = limpiar_texto(row.get("NOMBRE_CORRECTO"))
            if variante and nombre_correcto:
                homolog_generadores[variante] = nombre_correcto
        p(f"  Generadores homologados: {len(homolog_generadores)}")
    except Exception as e:
        p(f"  ⚠ No se pudo leer hoja generadores: {e}")

    return homolog_clientes, homolog_generadores


def cargar_lookups(rutas, p):
    p("\n── Cargando maestros compartidos desde OneDrive ──")

    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    df_t = df_t[["Transportista", "RUT"]].dropna(subset=["Transportista"])
    df_t["Transportista"] = df_t["Transportista"].astype(str).str.strip()
    df_t["RUT"] = df_t["RUT"].astype(str).str.strip()
    transportistas = dict(zip(df_t["Transportista"], df_t["RUT"]))
    p(f"  Transportistas: {len(transportistas)}")

    df_m = leer_excel(rutas["sinader"], p, sheet_name="Clasificación Residuos", header=1)
    df_m.columns = df_m.columns.astype(str).str.strip()
    df_m = df_m[["TIPO", "CÓDIGOS SINADER"]].dropna(subset=["TIPO"])
    df_m["TIPO"] = df_m["TIPO"].astype(str).str.strip()
    df_m["CÓDIGOS SINADER"] = df_m["CÓDIGOS SINADER"].astype(str).str.strip()
    df_m = df_m.drop_duplicates(subset=["TIPO"])
    materiales = dict(zip(df_m["TIPO"], df_m["CÓDIGOS SINADER"]))
    if "ASIMILABLE A DOMICILIARIO" not in materiales:
        materiales["ASIMILABLE A DOMICILIARIO"] = "20 03 01"
    p(f"  SINADER: {len(materiales)} tipos")

    df_d = leer_excel(rutas["destinatarios"], p, sheet_name="NACIONAL", header=1)
    df_d.columns = df_d.columns.astype(str).str.strip()
    cols = [
        "NOMBRE DE FANTASÍA", "COMUNA DE ESTABLECIMIENTO", "CÓDIGOS LER",
        "CÓDIGO ESTABLECIMIENTO SINADER", "RUT DESTINATARIO",
        "CÓDIGO DE TRATAMIENTO SINADER",
    ]
    df_d = df_d[cols].dropna(subset=["NOMBRE DE FANTASÍA"])
    for c in cols:
        df_d[c] = df_d[c].astype(str).str.strip()
    p(f"  Destinatarios: {len(df_d)}")

    return transportistas, materiales, df_d


# ══════════════════════════════════════════════════════════════════
# FUENTES
# ══════════════════════════════════════════════════════════════════
def preparar_control(nombre, df, mapa):
    out = pd.DataFrame(index=df.index)
    out["ORIGEN"] = nombre
    for salida, opciones in mapa.items():
        out[salida] = tomar_columna(df, opciones)
    out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce")
    out["AÑO"] = out["Fecha"].dt.year
    out["_desde"] = out["Fecha"] >= pd.Timestamp(FECHA_DESDE_SUR)
    return out


MAPA_ECOFIBRAS = {
    "Fecha":         ["FECHA", "Fecha"],
    "Cliente":       ["CLIENTE", "Cliente"],
    "RUT":           ["RUT CLIENTE", "RUT"],
    "Generador":     ["GENERADOR", "Generador"],
    "Transportista": ["TRANSPORTISTA", "Transportista", "GESTOR"],
    "Destino":       ["DESTINO", "Destino"],
    "TIPO":          ["TIPO", "TIPO RESIDUO"],
    "Movimiento":    ["Movimiento", "MOVIMIENTO"],
    "Peso neto":     ["PESO NETO KG", "Peso neto", "Peso neto (kg)"],
}


def cargar_fuentes_control(rutas, p):
    p("\nLeyendo fuentes Zona Sur...")

    coronel_raw       = leer_fuente(rutas["coronel"],       "INGRESOS",       p, headers_fallback=(8, 9, 7, 0))
    bo_trapen_raw     = leer_fuente(rutas["bo_trapen"],     "MOVIMIENTOS BO", p, headers_fallback=(8, 9, 7, 0))
    planta_trapen_raw = leer_fuente(rutas["planta_trapen"], "INGRESOS",       p, headers_fallback=(8, 9, 7, 0))
    san_joaquin_raw   = leer_fuente(rutas["san_joaquin"],   "INGRESOS",       p, headers_fallback=(8, 9, 7, 0))
    irar_raw          = leer_fuente(rutas["irar"],          "INGRESOS",       p, headers_fallback=(8, 9, 7, 0))
    bo_chillan_raw    = leer_fuente(rutas["bo_chillan"],    "TRASLADOS BO",   p, headers_fallback=(0, 1, 7, 8))
    temuco_raw        = leer_fuente(rutas["temuco"],        "BBDD",           p, headers_fallback=(7, 8, 0))

    if "bo_chiloe" in rutas and Path(rutas["bo_chiloe"]).exists():
        bo_chiloe_raw = leer_fuente(rutas["bo_chiloe"], "BBDD", p, headers_fallback=(6, 7, 0))
        # Misma regla que el consolidador: el contrato MOWI no entra.
        _col_cli_ch = next((c for c in ("Cliente", "CLIENTE") if c in bo_chiloe_raw.columns), None)
        if _col_cli_ch:
            _mask_mowi = bo_chiloe_raw[_col_cli_ch].apply(lambda x: "MOWI" in normalizar_texto(x))
            p(f"  BO Chiloé: exclusión MOWI → {int(_mask_mowi.sum())} filas fuera, {int((~_mask_mowi).sum())} al control")
            bo_chiloe_raw = bo_chiloe_raw.loc[~_mask_mowi].reset_index(drop=True)
    else:
        p("  ⚠ BBDD BO CHILOÉ.xlsx no encontrado — fuente omitida del control")
        bo_chiloe_raw = pd.DataFrame()

    fuentes = {
        "CORONEL": preparar_control("CORONEL", coronel_raw, MAPA_ECOFIBRAS),
        "BO_TRAPEN": preparar_control("BO_TRAPEN", bo_trapen_raw, {
            **MAPA_ECOFIBRAS,
            "TIPO": ["TIPO RESIDUO", "TIPO"],
            "Movimiento": ["MOVIMIENTO", "Movimiento"],
        }),
        "PLANTA_TRAPEN": preparar_control("PLANTA_TRAPEN", planta_trapen_raw, {
            **MAPA_ECOFIBRAS,
            "TIPO": ["TIPO RESIDUO", "TIPO"],
            "Movimiento": ["MOVIMIENTO", "Movimiento"],
        }),
        "SAN_JOAQUIN": preparar_control("SAN_JOAQUIN", san_joaquin_raw, {
            **MAPA_ECOFIBRAS,
            "TIPO": ["TIPO RESIDUO", "TIPO"],
            "Movimiento": ["MOVIMIENTO", "Movimiento"],
        }),
        "IRAR_LOS_ANGELES": preparar_control("IRAR_LOS_ANGELES", irar_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["RUT", "RUT CLIENTE"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["Destino", "DESTINO"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto", "PESO NETO KG", "Peso neto (kg)"],
        }),
        "BO_CHILLAN": preparar_control("BO_CHILLAN", bo_chillan_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["RUT", "RUT CLIENTE"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["Destino", "DESTINO"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto", "PESO NETO KG", "Peso neto (kg)"],
        }),
        "TEMUCO": preparar_control("TEMUCO", temuco_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["RUT", "RUT CLIENTE"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["Destino", "DESTINO"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto", "PESO NETO KG", "Peso neto (kg)"],
        }),
        "BO_CHILOE": preparar_control("BO_CHILOE", bo_chiloe_raw, {
            "Fecha":         ["Fecha", "FECHA"],
            "Cliente":       ["Cliente", "CLIENTE"],
            "RUT":           ["RUT", "RUT CLIENTE"],
            "Generador":     ["Generador", "GENERADOR"],
            "Transportista": ["Transportista", "TRANSPORTISTA"],
            "Destino":       ["Destino", "DESTINO"],
            "TIPO":          ["TIPO", "TIPO RESIDUO"],
            "Movimiento":    ["Movimiento", "MOVIMIENTO"],
            "Peso neto":     ["Peso neto [kg]", "Peso neto", "PESO NETO KG", "Peso neto (kg)"],
        }),
    }

    for nombre, df in fuentes.items():
        p(f"  {nombre:<18} {len(df):>6} filas | >= {FECHA_DESDE_SUR}: {df['_desde'].sum()}")

    return fuentes


# ══════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL — esta es la que llama la aplicación
# ══════════════════════════════════════════════════════════════════
def controlar(rutas, ruta_salida=None, mostrar=True):
    """Ejecuta el control de calidad de Sur sobre las fuentes indicadas.

    Devuelve un dict con las nueve hojas de control (c1, c2, c3, c_tipo,
    c_trans, c_vacios, c_mov, c_dest) más 'df_resumen' (filas leídas por
    fuente) y 'total'.
    """
    p = Registro(mostrar=mostrar)
    p("Iniciando control de calidad SUR...")

    verificar_rutas(rutas)

    fuentes = cargar_fuentes_control(rutas, p)
    _, materiales, _ = cargar_lookups(rutas, p)

    resumen = []
    partes = []

    for nombre, df in fuentes.items():
        resumen.append({
            "FUENTE": nombre,
            "FILAS_LEIDAS": len(df),
            f"FILAS_DESDE_{FECHA_DESDE_SUR[:7]}": int(df["_desde"].sum()),
            "FECHA_MIN": df["Fecha"].min(),
            "FECHA_MAX": df["Fecha"].max(),
        })
        partes.append(df[df["_desde"]].copy())

    df_resumen = pd.DataFrame(resumen)
    df_all = pd.concat(partes, ignore_index=True)

    # Movimientos según regla operacional
    df_all["Movimiento_norm"] = df_all["Movimiento"].apply(lambda x: normalizar_movimiento(x))
    df_all.loc[df_all["ORIGEN"].eq("TEMUCO"),           "Movimiento_norm"] = "Traslado"
    df_all.loc[df_all["ORIGEN"].eq("IRAR_LOS_ANGELES"), "Movimiento_norm"] = "Traslado"
    df_all.loc[df_all["ORIGEN"].eq("BO_TRAPEN"),        "Movimiento_norm"] = "Traslado"
    df_all.loc[df_all["ORIGEN"].eq("BO_CHILOE"),        "Movimiento_norm"] = "Traslado"
    df_all.loc[df_all["ORIGEN"].eq("PLANTA_TRAPEN"),    "Movimiento_norm"] = "Ingreso"

    # En hojas INGRESOS, movimiento vacío se interpreta como Ingreso para control.
    df_all.loc[
        df_all["ORIGEN"].isin(["CORONEL", "SAN_JOAQUIN"])
        & df_all["Movimiento_norm"].isna(),
        "Movimiento_norm"
    ] = "Ingreso"

    # Clientes / RUT
    clientes = df_all[["ORIGEN", "Cliente", "RUT"]].copy()
    clientes["CLIENTE_NORM"] = clientes["Cliente"].apply(normalizar_texto)
    clientes["RUT_NORM"] = clientes["RUT"].apply(lambda x: normalizar_texto(x).replace("-", ""))
    clientes = clientes[(clientes["CLIENTE_NORM"] != "") | (clientes["RUT_NORM"] != "")].drop_duplicates()

    rut_por_nombre = clientes.groupby("CLIENTE_NORM")["RUT_NORM"].nunique().reset_index(name="N_RUT")
    nombres_conf = rut_por_nombre[(rut_por_nombre["CLIENTE_NORM"] != "") & (rut_por_nombre["N_RUT"] > 1)]["CLIENTE_NORM"]
    c1 = clientes[clientes["CLIENTE_NORM"].isin(nombres_conf)].copy()
    if not c1.empty:
        c1.insert(0, "PROBLEMA", "Mismo cliente normalizado con RUT distinto")

    nombre_por_rut = clientes.groupby("RUT_NORM")["CLIENTE_NORM"].nunique().reset_index(name="N_CLIENTE")
    ruts_conf = nombre_por_rut[(nombre_por_rut["RUT_NORM"] != "") & (nombre_por_rut["N_CLIENTE"] > 1)]["RUT_NORM"]
    c2 = clientes[clientes["RUT_NORM"].isin(ruts_conf)].copy()
    if not c2.empty:
        c2.insert(0, "PROBLEMA", "Mismo RUT con nombre distinto")

    # Tipo sin SINADER
    tipos = df_all[["ORIGEN", "TIPO"]].drop_duplicates().copy()
    tipos["TIPO_LIMPIO"] = tipos["TIPO"].astype(str).str.strip()
    c_tipo = tipos[
        ~tipos["TIPO_LIMPIO"].isin(set(materiales.keys()))
        & (~tipos["TIPO_LIMPIO"].str.lower().isin(["nan", "none", ""]))
    ].copy()
    if not c_tipo.empty:
        c_tipo.insert(0, "PROBLEMA", "TIPO sin código SINADER")

    # Transportistas sin RUT
    df_t = leer_excel(rutas["transportistas"], p, sheet_name="Hoja1", header=0)
    df_t.columns = df_t.columns.astype(str).str.strip()
    trans_ok = set(df_t["Transportista"].dropna().astype(str).str.strip())

    trans = df_all[["ORIGEN", "Transportista"]].drop_duplicates().copy()
    trans["Transportista"] = trans["Transportista"].apply(limpiar_texto)
    c_trans = trans[(trans["Transportista"].notna()) & (~trans["Transportista"].isin(trans_ok))].copy()
    if not c_trans.empty:
        c_trans.insert(0, "PROBLEMA", "Transportista sin RUT en maestro Transportistas")

    # Vacíos críticos
    criticas = ["Fecha", "Cliente", "RUT", "Generador", "TIPO", "Movimiento_norm", "Peso neto"]
    filas_vacios = []
    for col in criticas:
        serie = df_all[col] if col in df_all.columns else pd.Series([None] * len(df_all))
        vacio = serie.isna() | serie.astype(str).str.strip().str.lower().isin(["", "nan", "none", "nat", "#n/a"])
        if vacio.any():
            por_origen = df_all.loc[vacio].groupby("ORIGEN").size().reset_index(name="CANTIDAD")
            for _, r in por_origen.iterrows():
                filas_vacios.append({
                    "PROBLEMA": "Columna crítica vacía",
                    "COLUMNA": col,
                    "ORIGEN": r["ORIGEN"],
                    "CANTIDAD": int(r["CANTIDAD"]),
                })
    c_vacios = pd.DataFrame(filas_vacios)

    # Destinos vacíos solo importan cuando movimiento final será Traslado o fuente Temuco/IRAR
    destinos_revisar = df_all.copy()
    necesita_destino_archivo = (
        destinos_revisar["ORIGEN"].isin(["TEMUCO", "IRAR_LOS_ANGELES", "BO_CHILOE"])
        | (
            destinos_revisar["ORIGEN"].isin(["CORONEL", "BO_TRAPEN", "SAN_JOAQUIN"])
            & destinos_revisar["Movimiento_norm"].eq("Traslado")
        )
    )
    destino_vacio = destinos_revisar["Destino"].isna() | destinos_revisar["Destino"].astype(str).str.strip().str.lower().isin(["", "nan", "none", "nat"])
    c_dest = destinos_revisar.loc[necesita_destino_archivo & destino_vacio, ["ORIGEN", "Fecha", "Cliente", "Movimiento_norm", "Destino"]].copy()
    if not c_dest.empty:
        c_dest.insert(0, "PROBLEMA", "Destino vacío donde corresponde usar columna Destino")

    # C3: Generadores con variantes de nombre entre bases
    generadores = df_all[["ORIGEN", "Generador"]].copy()
    generadores["GEN_NORM"] = generadores["Generador"].apply(normalizar_texto)
    generadores = generadores[generadores["GEN_NORM"] != ""].drop_duplicates()

    nombres_por_gen = generadores.groupby("GEN_NORM")["Generador"].nunique().reset_index(name="N_NOMBRE")
    gens_conf = nombres_por_gen[nombres_por_gen["N_NOMBRE"] > 1]["GEN_NORM"]
    c3 = generadores[generadores["GEN_NORM"].isin(gens_conf)].copy()
    if not c3.empty:
        c3.insert(0, "PROBLEMA", "Generador con variantes de nombre entre bases")

    # Nombres para poblar HOMOLOGACION.xlsx
    cli_homolog = (
        clientes[["ORIGEN", "Cliente", "RUT"]]
        .drop_duplicates().sort_values(["ORIGEN", "Cliente"])
        .rename(columns={"ORIGEN": "PLANTA", "Cliente": "NOMBRE_VARIANTE"})
        .reset_index(drop=True)
    )
    cli_homolog["NOMBRE_CORRECTO"] = ""
    cli_homolog["RUT_CORRECTO"] = ""

    gen_homolog = (
        generadores[["ORIGEN", "Generador"]]
        .drop_duplicates().sort_values(["ORIGEN", "Generador"])
        .rename(columns={"ORIGEN": "PLANTA", "Generador": "NOMBRE_VARIANTE"})
        .reset_index(drop=True)
    )
    gen_homolog["NOMBRE_CORRECTO"] = ""

    # Movimientos raros
    mov_ok = {"Ingreso", "Traslado"}
    c_mov = df_all[
        df_all["Movimiento_norm"].isna() | (~df_all["Movimiento_norm"].isin(mov_ok))
    ][["ORIGEN", "Fecha", "Cliente", "Movimiento", "Movimiento_norm"]].drop_duplicates().copy()
    if not c_mov.empty:
        c_mov.insert(0, "PROBLEMA", "Movimiento vacío o no reconocido")

    total = len(c1) + len(c2) + len(c3) + len(c_tipo) + len(c_trans) + len(c_vacios) + len(c_mov) + len(c_dest)

    res = {
        "df_resumen": df_resumen,
        "c1": c1, "c2": c2, "c3": c3,
        "c_tipo": c_tipo, "c_trans": c_trans, "c_vacios": c_vacios,
        "c_mov": c_mov, "c_dest": c_dest,
        "cli_homolog": cli_homolog, "gen_homolog": gen_homolog,
        "total": total,
    }

    if ruta_salida:
        escribir_excel(res, ruta_salida, p)

    res["log"] = p.texto()
    return res


def escribir_excel(res, ruta_salida, p):
    p(f"\nGuardando {Path(ruta_salida).name}...")

    def hoja(writer, df, nombre, msg_ok):
        out = df if not df.empty else pd.DataFrame({"resultado": [msg_ok]})
        out.to_excel(writer, sheet_name=nombre, index=False)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        res["df_resumen"].to_excel(writer, sheet_name="RESUMEN", index=False)
        hoja(writer, res["c1"], "C1_CLIENTE_RUT", "Sin conflictos")
        hoja(writer, res["c2"], "C2_RUT_CLIENTE", "Sin conflictos")
        hoja(writer, res["c3"], "C3_GENERADORES", "Sin variantes")
        res["cli_homolog"].to_excel(writer, sheet_name="HOMOLOG_clientes", index=False)
        res["gen_homolog"].to_excel(writer, sheet_name="HOMOLOG_generadores", index=False)
        hoja(writer, res["c_tipo"], "TIPO_SIN_SINADER", "Todos los TIPO tienen SINADER")
        hoja(writer, res["c_trans"], "TRANSPORT_SIN_RUT", "Todos los transportistas tienen RUT")
        hoja(writer, res["c_vacios"], "VACIOS_CRITICOS", "Sin vacíos críticos")
        hoja(writer, res["c_mov"], "MOVIMIENTOS_REVISAR", "Movimientos OK")
        hoja(writer, res["c_dest"], "DESTINOS_REVISAR", "Sin destinos vacíos críticos")

    p("✓ Control de calidad SUR generado")
    p(f"  Archivo: {ruta_salida}")


# ══════════════════════════════════════════════════════════════════
# USO DESDE LA TERMINAL
# ══════════════════════════════════════════════════════════════════
def _main():
    argumentos = [a for a in sys.argv[1:] if not a.startswith("--")]

    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    try:
        from config import carpetas, ConfiguracionFaltante
    except ImportError:
        print("\n✗ No se encontró config.py. Debe estar en la carpeta src/.\n")
        sys.exit(2)

    try:
        if argumentos:
            rutas = rutas_desde_carpeta(argumentos[0])
            salida = Path(argumentos[0]) / "CONTROL_CALIDAD_SUR.xlsx"
        else:
            rutas = rutas_desde_config()
            salida = carpetas()["sur"] / "CONTROL_CALIDAD_SUR.xlsx"
    except ConfiguracionFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)
    except ArchivoFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    # Sello de fecha/hora para no pisar corridas anteriores (igual que RM)
    sello = datetime.now().strftime("%Y-%m-%d_%H%M")
    salida = salida.with_name(f"{salida.stem}_{sello}{salida.suffix}")

    try:
        res = controlar(rutas, ruta_salida=salida)
    except ArchivoFaltante as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)
    except ValueError as e:
        print(f"\n✗ {e}\n")
        sys.exit(2)

    sys.exit(0 if res["total"] == 0 else 1)


if __name__ == "__main__":
    _main()
